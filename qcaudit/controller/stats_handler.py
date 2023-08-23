from . import *
from qcaudit.env_config.pre_req import *
from qcaudit.common.const import *
from .qc_util import QCUtil
from qcaudit.service.protomarshaler import *
import threading, time, traceback, arrow
from datetime import timedelta, datetime
from multiprocessing import Process
from qcaudit.config import Config
from openpyxl import Workbook
from qcaudit.domain.stats.statsrepository import StatsRepository
from qcaudit.utils.towebconfig import *
from qcaudit.utils.bidataprocess import *
from qcaudit.domain.problem.statsreq import GetProblemStatsRequest
from openpyxl.styles import Font
from qcaudit.domain.stats.archived_quality import ArchivedQualityStats


class StatsCaseRatio(MyResource):

    @pre_request(request, StatsCaseRatioReq)
    def post(self):
        """
        获取病历归档率、完成率、退回率、等级占比
        :param request:
        :param context:
        :return:
        """
        response = {"caseRateInfo": [], "levelRatioInfo": {}}
        startTime, endTime = QCUtil.get_start_end_time(request)
        if not startTime or not endTime:
            return get_error_resp("time error.")

        with self.context.app.mysqlConnection.session() as cursor:
            stats_name_list = ['申请归档病历', '完成归档病历', '退回整改病历']
            for stats_name in stats_name_list:
                call_approve_rate_sql = '''call pr_case_qc('%s','%s','%s', '')''' % (stats_name, startTime, endTime)
                query = cursor.execute(call_approve_rate_sql)
                self.logger.info("StatsCaseRatio, execute call_approve_rate_sql: %s", call_approve_rate_sql)
                queryset = query.fetchone()
                approve_case_count = queryset[0] or 0
                sameCompareRate = queryset[1] or 0.00
                chainCompareRate = queryset[2] or 0.00
                approve_rate = queryset[3] or 0.00

                proto_case_rate = {}  # response.caseRateInfo.add()
                proto_case_rate["statsName"] = stats_name
                proto_case_rate["count"] = int(approve_case_count)
                proto_case_rate["rate"] = QCUtil.keepOne(approve_rate)
                proto_case_rate["sameCompareRate"] = QCUtil.keepOne(sameCompareRate)
                proto_case_rate["chainCompareRate"] = QCUtil.keepOne(chainCompareRate)
                response["caseRateInfo"].append(proto_case_rate)

            stats_name = '病历等级占比分析'
            call_level_rate_sql = '''call pr_case_qc('%s','%s','%s', '')''' % (stats_name, startTime, endTime)
            self.logger.info("StatsCaseRatio, execute call_level_rate_sql: %s", call_level_rate_sql)
            query = cursor.execute(call_level_rate_sql)
            queryset = query.fetchall()
            total = 0
            first = 0
            second = 0
            third = 0
            for item in queryset:
                count = item[1] or 0
                if item[0] == "甲":
                    first += count
                if item[0] == "乙":
                    second += count
                if item[0] == "丙":
                    third += count
                total += count

            proto_level_ratio = response["levelRatioInfo"]
            proto_level_ratio["first"] = QCUtil.keepOne(first / total * 100) if total else QCUtil.keepOne(0)
            proto_level_ratio["second"] = QCUtil.keepOne(second / total * 100) if total else QCUtil.keepOne(0)
            proto_level_ratio["third"] = QCUtil.keepOne(third / total * 100) if total else QCUtil.keepOne(0)

        return response


class StatsDepartmentScore(MyResource):

    @pre_request(request, StatsCaseRatioReq)
    def post(self):
        """
        获取各科室质控成绩统计
        :param request:
        :param context:
        :return:
        """
        response = {"items": []}
        startTime, endTime = QCUtil.get_start_end_time(request)
        if not startTime or not endTime:
            return get_error_resp("time error.")

        with self.context.app.mysqlConnection.session() as cursor:
            stats_name = "各科室质控成绩统计"
            call_department_score_sql = '''call pr_case_qc('%s','%s','%s', '')''' % (stats_name, startTime, endTime)
            query = cursor.execute(call_department_score_sql)
            self.logger.info("StatsDepartmentScore, execute call_department_score_sql: %s", call_department_score_sql)
            queryset = query.fetchall()

            for item in queryset:
                if not item[0]:
                    continue
                protoItem = {}  # response.items.add()
                protoItem["departmentName"] = item[0]
                protoItem["firstNum"] = int(item[1] or 0)
                protoItem["secondNum"] = int(item[2] or 0)
                protoItem["thirdNum"] = int(item[3] or 0)
                protoItem["approveRatio"] = QCUtil.keepOne(item[4] or 0)
                protoItem["returnRatio"] = QCUtil.keepOne(item[5] or 0)
                response["items"].append(protoItem)

        return response


class StatsCaseTarget(MyResource):

    @pre_request(request, StatsCaseRatioReq)
    def post(self):
        """
        获取病案指标统计
        :param request:
        :param context:
        :return:
        """
        response = {"data": []}
        startTime, endTime = QCUtil.get_start_end_time(request)
        targetName = request.targetName or ""
        if not startTime or not endTime:
            return get_error_resp("time error.")
        target_list = targetName.split(",")

        with self.context.app.mysqlConnection.session() as cursor:
            stats_name = "病案质量管理指标分析"
            call_case_target_sql = '''call pr_case_qc('%s','%s','%s', '')''' % (stats_name, startTime, endTime)
            query = cursor.execute(call_case_target_sql)
            self.logger.info("StatsCaseTarget, execute call_case_target_sql: %s", call_case_target_sql)
            queryset = query.fetchall()

            parent_target_dict = {}
            for item in queryset:
                parent_target = item[0]
                son_target = item[1]
                rate = QCUtil.keepOne(item[2] or 0)
                sameCompareRate = QCUtil.keepOne(item[3] or 0)
                chainCompareRate = QCUtil.keepOne(item[4] or 0)
                if targetName and parent_target not in target_list:
                    self.logger.info("StatsCaseTarget, target_list is %s, parent_target is %s, not need show",
                                     target_list, parent_target)
                    continue
                if not parent_target_dict.get(parent_target, ""):
                    protoItem = {"targetData": []}  # response.data.add()
                    parent_target_dict[parent_target] = protoItem
                    protoItem["parentTargetName"] = parent_target
                else:
                    protoItem = parent_target_dict[parent_target]

                protoTarget = {}  # protoItem.targetData.add()
                protoTarget["sonTargetName"] = son_target
                protoTarget["rate"] = rate
                protoTarget["sameCompareRate"] = sameCompareRate
                protoTarget["chainCompareRate"] = chainCompareRate
                protoItem["targetData"].append(protoTarget)

                response["data"].append(protoItem)

        return response


class StatsCaseDefectRate(MyResource):

    @pre_request(request, StatsCaseRatioReq)
    def post(self):
        """
        获取病历平均缺陷率
        :param request:
        :param context:
        :return:
        """
        response = {"items": []}
        startTime, endTime = QCUtil.get_start_end_time(request)
        if not startTime or not endTime:
            return get_error_resp("time error.")
        today = datetime.now().strftime("%Y-%m-%d")
        month = today[:-3]

        with self.context.app.mysqlConnection.session() as cursor:
            stats_name = "病历平均缺陷统计月分析" if request.timeType == "month" else "病历平均缺陷统计年分析"
            call_case_defect_sql = '''call pr_case_qc('%s','%s','%s', '')''' % (stats_name, startTime, endTime)
            query = cursor.execute(call_case_defect_sql)
            self.logger.info("StatsCaseDefectRate, execute call_case_defect_sql: %s", call_case_defect_sql)
            queryset = query.fetchall()

            for item in queryset:
                if request.timeType == "month":
                    if str(item[0][:-3]) > month or (str(item[0][:-3]) == month and str(item[0]) > today):
                        # 大于当月或等于当月大于当天
                        break
                else:
                    if str(item[0]) > month:
                        break

                protoItem = {}  # response.items.add()
                date = int(item[0][-2:])
                protoItem["xName"] = "{}号".format(date) if request.timeType == "month" else "{}月".format(date)
                protoItem["defectRate"] = QCUtil.keepOne(item[1] or 0)
                protoItem["firstPageDefectRate"] = QCUtil.keepOne(item[2] or 0)
                response["items"].append(protoItem)

        return response


class StatsCaseDefectCount(MyResource):

    @pre_request(request, StatsCaseRatioReq)
    def post(self):
        """
        获取缺陷数量统计
        :param request:
        :param context:
        :return:
        """
        response = {"items": []}
        startTime, endTime = QCUtil.get_start_end_time(request)
        if not startTime or not endTime:
            return get_error_resp("time error.")
        isFirstPage = request.isFirstPage

        with self.context.app.mysqlConnection.session() as cursor:
            stats_name = "常见缺陷发生数量统计"
            call_case_defect_count_sql = '''call pr_case_qc('%s','%s','%s', '')''' % (stats_name, startTime, endTime)
            query = cursor.execute(call_case_defect_count_sql)
            self.logger.info("StatsCaseDefectCount, execute call_case_defect_count_sql: %s", call_case_defect_count_sql)
            queryset = query.fetchall()
            if queryset:
                for item in queryset:
                    protoItem = {}
                    if isFirstPage == 0:
                        protoItem = {}  # response.items.add()
                        protoItem["defectName"] = item[0] or ""
                        protoItem["nowCount"] = int(item[1] or 0)
                        protoItem["pastCount"] = int(item[4] or 0)
                    elif isFirstPage == 1:
                        if item[2] != 0:
                            protoItem = {}  # response.items.add()
                            protoItem["defectName"] = item[0] or ""
                            protoItem["nowCount"] = int(item[2] or 0)
                            protoItem["pastCount"] = int(item[5] or 0)
                    elif isFirstPage == 2:
                        if item[3] != 0:
                            protoItem = {}  # response.items.add()
                            protoItem["defectName"] = item[0] or ""
                            protoItem["nowCount"] = int(item[3] or 0)
                            protoItem["pastCount"] = int(item[6] or 0)
                    if protoItem:
                        response["items"].append(protoItem)

        return response


class StatsFlagCaseDefectList(MyResource):
    
    _FLAG_CASE_DEFECT_DICT = {}

    @pre_request(request, StatsCaseRatioReq)
    def post(self):
        """
        获取重点病历缺陷分析
        :param request:
        :param context:
        :return:
        """
        response = {"items": []}
        startTime, endTime = QCUtil.get_start_end_time(request)
        if not startTime or not endTime:
            return get_error_resp("time error.")
        sortTags = request.sortTags or []

        self.update_tag_status_sort(sortTags)
        self.query_flag_case_defect_by_thread(sortTags, startTime, endTime)

        for case_flag in sortTags:
            if case_flag.status != 1 or case_flag.name not in self._FLAG_CASE_DEFECT_DICT:
                continue
            item = self._FLAG_CASE_DEFECT_DICT[case_flag.name]
            protoItem = {}  # response.items.add()
            protoItem["caseFlag"] = case_flag.name
            protoItem["caseCount"] = int(item["caseCount"] or 0)
            protoItem["defectCount"] = int(item["defectCount"] or 0)
            protoItem["defectRate"] = QCUtil.keepOne(item["defectRate"] or 0)
            protoItem["sameCompareRate"] = QCUtil.keepOne(item["sameCompareRate"] or 0)
            protoItem["chainCompareRate"] = QCUtil.keepOne(item["chainCompareRate"] or 0)
            response["items"].append(protoItem)

        return response

    def update_tag_status_sort(self, sortTags):
        """
        更新重点病历标签状态和排序方式
        :return:
        """
        query_old_tags_sql = '''select id, status, no from tags_qc'''
        self.logger.info("update_tag_status_sort, query_old_tags_sql: %s", query_old_tags_sql)
        with self.context.app.mysqlConnection.session() as cursor:
            query = cursor.execute(query_old_tags_sql)
            data = query.fetchall()
        old_tags = {item[0]: [int(item[1]), int(item[2])] for item in data}
        diff_tags = {item.id: [item.status, item.orderNo] for item in sortTags
                     if item.status != old_tags[item.id][0] or item.orderNo != old_tags[item.id][1]}

        self.logger.info("update_tag_status_sort, diff_tags: %s", diff_tags)
        if diff_tags:
            for tag_id in diff_tags:
                new_status = diff_tags[tag_id][0]
                new_orderNo = diff_tags[tag_id][1]
                update_sql = '''update tags_qc set status = "%s", no = "%s" where id = "%s"''' % (
                    new_status, new_orderNo, tag_id)
                cursor.execute(update_sql)
            cursor.commit()

    def query_flag_case_defect_by_thread(self, sortTags, startTime, endTime):
        """
        启动多线程
        :return:
        """
        threads = []
        for tag in sortTags:
            if tag.status == 1:
                t = threading.Thread(target=self.thread_call_get_data, args=(startTime, endTime, tag.name))
                threads.append(t)

        for t in threads:
            t.setDaemon(True)
            t.start()
        for t in threads:
            t.join()

    def thread_call_get_data(self, startTime, endTime, name):
        """
        在存储过程获取病历缺陷数据线程函数
        :return:
        """
        with self.context.app.mysqlConnection.session() as cursor:
            call_flag_case_sql = '''call pr_case_qc('重点病历缺陷分析','{}','{}', '{}')'''
            query = cursor.execute(call_flag_case_sql.format(startTime, endTime, name))
            queryset = query.fetchall()
            if queryset:
                queryset = queryset[0]
                self._FLAG_CASE_DEFECT_DICT[queryset[0]] = {"caseCount": queryset[1], "defectCount": queryset[2],
                                                            "defectRate": queryset[3], "sameCompareRate": queryset[5],
                                                            "chainCompareRate": queryset[4]}


class GetStatsDataUpdateStatus(MyResource):

    def get(self):
        """
        获取统计数据更新状态信息(首页评分,归档率等统计)
        :param request:
        :param context:
        :return:
        """
        response = {}
        result = QCUtil.getStatsUpdateState(self.context.app)
        response["status"] = result.get("status") if result.get("status") else 0
        updateDatetime = result.get("updatetime") + timedelta(hours=8)
        self.logger.info(updateDatetime)
        response["lastUpdateTime"] = updateDatetime.strftime('%Y-%m-%d %H:%M:%S') if result.get("updatetime") else ""
        return response


class StatsDataUpdate(MyResource):

    def get(self):
        """
        更新统计数据接口(首页评分,归档率等统计)
        :param request:
        :param context:
        :return:
        """
        response = {}

        result = QCUtil.getStatsUpdateState(self.context.app)
        if result.get("status") == 1 or result.get("updatetime") > (datetime.now() - timedelta(minutes=5)):
            return get_error_resp("数据更新频率太高")

        with self.context.app.mysqlConnection.session() as cursor:
            update_sql = "update case_updatestatus set updatestatus = 1, updatetime = Now();"
            ret = cursor.execute(update_sql)
            if ret:
                self.logger.info(ret)
                response["isSuccess"] = True
                self.logger.info("start finishDataUpdate time is : %s " % time.time())
                p = Process(target=self.finishDataUpdate, args=(cursor, ))
                p.daemon = True
                p.start()
                self.logger.info("finishDataUpdate time is : %s " % time.time())
        return response


class GetStatsTableUpdateStatus(MyResource):

    @pre_request(request, ["type"])
    def get(self):
        """
        统计数据更新状态信息
        :param request:
        :param context:
        :return:
        """
        response = {}

        result = QCUtil.getLastUpdateTime(self.app, request.type)
        response["status"] = result.get("status", 0)
        updateDatetime = result.get("updateTime", "")
        self.logger.info("lastUpdateTime: %s", updateDatetime)
        response["lastUpdateTime"] = updateDatetime.strftime('%Y-%m-%d %H:%M:%S') if updateDatetime else ""
        return response


class StatsTableUpdate(MyResource):

    @pre_request(request, ["type"])
    def get(self):
        """
        数据更新
        :param request:
        :param context:
        :return:
        """
        result = QCUtil.getLastUpdateTime(self.app, request.type)
        can_update_time = datetime.now() + timedelta(hours=8) - timedelta(minutes=5)
        self.logger.info("StatsTableUpdate, can_update_time: %s", can_update_time)
        if result.get("updateTime", "") and result.get("updateTime") > can_update_time:
            return get_error_resp("数据更新频率太高")

        with self.context.app.mysqlConnection.session() as cursor:
            table = self.get_update_status_table(request.type)
            update_this_time_sql = "update %s set updatestatus = 1, updatetime = Now();" % table
            ret = cursor.execute(update_this_time_sql)
            cursor.commit()
            self.logger.info("StatsDataUpdate, update_this_time_sql: %s", update_this_time_sql)
            if ret:
                try:
                    p = threading.Thread(target=self.finishTableUpdate, args=(request.type, ))
                    p.start()
                except Exception:
                    self.logger.error("StatsDataUpdate, finishTableUpdate, error: %s", traceback.format_exc())
                    return get_error_resp("更新失败")

        return g.result


class GetHospitalArchivingRate(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def post(self):
        """
        统计全院归档率
        :param request:
        :param context:
        :return:
        """
        response = {"data": []}

        branch = request.branch or "全部"
        args = [request.startTime, request.endTime, 'hospital', branch, '全部', "全部", ""]
        with self.context.app.mysqlConnection.session() as cursor:
            call_proc_sql = """call pr_case_archiverate('%s', '%s', '%s', '%s', '%s', '%s', '%s')""" % tuple(args)
            query = cursor.execute(call_proc_sql)
            queryset = query.fetchall()
            for item in queryset:
                protoItem = {}  # response.data.add()
                protoItem["branch"] = item[0] or ""
                protoItem["dischargeCount"] = int(item[1] or 0)
                protoItem["applyCount"] = int(item[2] or 0)
                protoItem["refusedCount"] = int(item[3] or 0)
                protoItem["archivedCount"] = int(item[4] or 0)
                protoItem["unreviewCount"] = int(item[5] or 0)
                protoItem["archivingRate"] = float(item[6] or 0)
                protoItem["refusedRate"] = float(item[7] or 0)
                protoItem["finishRate"] = float(item[8] or 0)
                protoItem["imperfect"] = float(item[9] or 0)
                protoItem["secondApply"] = int(item[10] or 0)
                protoItem["secondNoApply"] = int(item[11] or 0)
                protoItem["secondArchivingRate"] = float(item[12] or 0)
                protoItem["thirdApply"] = int(item[13] or 0)
                protoItem["thirdNoApply"] = int(item[14] or 0)
                protoItem["thirdArchivingRate"] = float(item[15] or 0)
                protoItem["seventhApply"] = int(item[16] or 0)
                protoItem["seventhNoApply"] = int(item[17] or 0)
                protoItem["seventhArchivingRate"] = float(item[18] or 0)

                protoItem["timelyRate"] = str(item[19] if item[19] else 0)
                protoItem["fixRate"] = str(item[20] if item[20] else 0)
                response["data"].append(protoItem)

        return response


class GetDepartmentArchivingRate(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def post(self):
        """
        统计科室归档率
        :param request:
        :param context:
        :return:
        """
        response = {"data": []}
        sizeCount = 0

        branch = request.branch or "全部"
        department = request.department or "全部"
        args = [request.startTime, request.endTime, 'dept', branch, department, "全部", ""]
        is_need_summary = int(self.context.app.config.get(Config.QC_STATS_DEPT_ARCHIVE) or 0)

        with self.context.app.mysqlConnection.session() as cursor:
            call_proc_sql = """call pr_case_archiverate('%s', '%s', '%s', '%s', '%s', '%s', '%s')""" % tuple(args)
            query = cursor.execute(call_proc_sql)
            queryset = query.fetchall()
            # 科室归档率排序
            sortedResult = QCUtil.sortDepartmentArchivingRateResult(queryset)
            for item in sortedResult:
                if is_need_summary != 1 and "汇总" in (item[0] or ""):
                    continue
                sizeCount += 1
                protoItem = {}  # response.data.add()
                protoItem["department"] = item[0] or ""
                protoItem["dischargeCount"] = int(item[1]) if item[1] else 0
                protoItem["applyCount"] = int(item[2]) if item[2] else 0
                protoItem["refusedCount"] = int(item[3]) if item[3] else 0
                protoItem["archivedCount"] = int(item[4]) if item[4] else 0
                protoItem["unreviewCount"] = int(item[5]) if item[5] else 0
                protoItem["archivingRate"] = float(item[6]) if item[6] else 0
                protoItem["refusedRate"] = float(item[7]) if item[7] else 0
                protoItem["finishRate"] = float(item[8]) if item[8] else 0
                protoItem["imperfect"] = float(item[9]) if item[9] else 0

                protoItem["secondApply"] = int(item[10]) if item[10] else 0
                protoItem["secondNoApply"] = int(item[11]) if item[11] else 0
                protoItem["secondArchivingRate"] = float(item[12]) if item[12] else 0
                protoItem["thirdApply"] = int(item[13]) if item[13] else 0
                protoItem["thirdNoApply"] = int(item[14]) if item[14] else 0
                protoItem["thirdArchivingRate"] = float(item[15]) if item[15] else 0
                protoItem["seventhApply"] = int(item[16]) if item[16] else 0
                protoItem["seventhNoApply"] = int(item[17]) if item[17] else 0
                protoItem["seventhArchivingRate"] = float(item[18]) if item[18] else 0

                protoItem["timelyRate"] = str(item[19] if item[19] else 0)
                protoItem["fixRate"] = str(item[20] if item[20] else 0)
                response["data"].append(protoItem)

        return response


class ExportDepartmentArchivingRate(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def post(self):
        """
        导出科室归档率
        :param request:
        :param context:
        :return:
        """
        response = {}

        branch = request.branch or "全部"
        department = request.department or "全部"
        args = [request.startTime, request.endTime, 'dept', branch, department, '全部', ""]

        excelTitle = "{0}至{1}各科室病历完成情况".format(request.startTime, request.endTime)
        exportName = "科室归档率{0}至{1}.xlsx".format(request.startTime, request.endTime)
        is_need_summary = int(self.context.app.config.get(Config.QC_STATS_DEPT_ARCHIVE) or 0)
        with self.context.app.mysqlConnection.session() as cursor:
            fileId = QCUtil.getFileId()
            filename = "%s.xlsx" % fileId
            fullname = os.path.join(export_path, filename)
            wb = Workbook()
            call_proc_sql = """call pr_case_archiverate('%s', '%s', '%s', '%s', '%s', '%s', '%s')""" % tuple(args)
            self.logger.info("call_proc_sql: %s", json.dumps(call_proc_sql, ensure_ascii=False))
            query = cursor.execute(call_proc_sql)
            queryset = query.fetchall()
            retCols = query.keys()
            if queryset:
                sortedResult = QCUtil.sortDepartmentArchivingRateResult(queryset)
                QCUtil.generateWorkSheet(sortedResult, wb, retCols, title=excelTitle, exportType="department", is_need_summary=is_need_summary)
            wb.save(fullname)
        response["id"] = str(fileId)
        response["fileName"] = exportName
        return response


class GetMonthArchivingRate(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def post(self):
        """
        统计科室归档率-按月份统计
        :param request:
        :param context:
        :return:
        """
        response = {"data": []}
        endTime = request.endTime or datetime.now().strftime("%Y-%m-%d")
        startTime = StatsRepository.get_year_start_time(endTime)
        args = [startTime, endTime, "month", request.branch or "全部", request.department or "全部", "全部", "全部"]

        with self.context.app.mysqlConnection.session() as session:
            call_proc_sql = """call pr_case_archiverate('%s', '%s', '%s', '%s', '%s', '%s', '%s')""" % tuple(args)
            query = session.execute(call_proc_sql)
            queryset = query.fetchall()
            for item in queryset:
                protoItem = {}  # response.data.add()
                protoItem["month"] = item[0]
                protoItem["archivingRate"] = QCUtil.keep_one(item[1])
                protoItem["archivingRate2"] = QCUtil.keep_one(item[2])
                protoItem["archivingRate3"] = QCUtil.keep_one(item[3])
                protoItem["archivingRate7"] = QCUtil.keep_one(item[4])
                response["data"].append(protoItem)

        return response


class GetBranchTimelinessRate(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def post(self):
        """
        全院病案指标-病历书写时效性、诊疗行为符合率等
        :param request:
        :param context:
        :return:
        """
        yaml_dict = {
            1: BRANCH_TIMELINESS_RATE_YAML.format(branch='院区'),
            2: BRANCH_TIMELINESS_RATE_YAML.format(branch='科室'),
            3: DOCTOR_TIMELINESS_RATE_YAML,
            4: BRANCH_TIMELINESS_RATE_YAML.format(branch='病区')
        }
        target_dict = {
            1: '全院病案指标',
            2: '科室病案指标',
            3: '医生病案指标',
            4: '病区病案指标'
        }
        target = target_dict.get(request.statusType,'')
        if not target:
            return get_error_resp("参数错误")
        response = {"headers": [], "data": [], "pageInfo": {}}
        args = [target, request.startTime or "2000-01-01", request.endTime or "2030-12-31", request.branch or "全部",
                request.department or request.ward or "全部", request.attend or "全部"]
        start = request.start or 0
        size = request.size or 10
        end = start + size

        with self.context.app.mysqlConnection.session() as session:
            call_proc_sql = """call pr_case_medical_index('%s','%s', '%s', '%s', '%s', '%s')""" % tuple(args)
            self.logger.info("GetBranchTimelinessRate, call_proc_sql: %s", call_proc_sql)
            query = session.execute(call_proc_sql)
            queryset = query.fetchall()
            total = sum([1 for item in queryset if item[0]])
            retCols = query.keys()[:-1]
            row_data = []
            for item in queryset[start: end]:
                if not item[0]:
                    continue
                tmp = {}
                for index in range(len(retCols)):
                    tmp[retCols[index]] = str(item[index]).replace("--", "")
                row_data.append(tmp)
            yaml_str = yaml_dict.get(request.statusType, None)
            if not yaml_str:
                return
            yaml_str = StatsRepository.get_branch_timeliness_yaml(yaml_str, self.context.app.config.get(Config.QC_STATS_BRANCH_TARGET_FIELD) or "")
            cfg = BIFormConfig.fromYaml(yaml_str)
            processor = BIDataProcess(cfg, row_data)
            header, result = processor.toWeb()
            QCUtil.format_header_data(response, header, result, first_title_dict=BRANCH_TIMELINESS_RATE_TARGET_FIRST_NAME_DICT)
            response["pageInfo"]["total"] = total
            response["pageInfo"]["size"] = size
            response["pageInfo"]["start"] = start

        return response


class GetBranchTimelinessRateExport(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def post(self):
        """
        全院病案指标-病历书写时效性、诊疗行为符合率等-导出
        :param request:
        :param context:
        :return:
        """
        yaml_dict = {
            1: BRANCH_TIMELINESS_RATE_YAML.format(branch='院区'),
            2: BRANCH_TIMELINESS_RATE_YAML.format(branch='科室'),
            3: DOCTOR_TIMELINESS_RATE_YAML,
            4: BRANCH_TIMELINESS_RATE_YAML.format(branch='病区')
        }
        target_dict = {
            1: '全院病案指标',
            2: '科室病案指标',
            3: '医生病案指标',
            4: '病区病案指标'
        }
        response = {}
        target = target_dict.get(request.statusType, '')
        if not target:
            return
        args = [target, request.startTime or "2000-01-01", request.endTime or "2030-12-31", request.branch or "全部",
                request.department or request.ward or "全部", request.attend or "全部"]

        with self.context.app.mysqlConnection.session() as session:
            call_proc_sql = """call pr_case_medical_index('%s','%s', '%s', '%s', '%s', '%s')""" % tuple(args)
            query = session.execute(call_proc_sql)
            self.logger.info("GetBranchTimelinessRateExport, call_proc_sql: %s", call_proc_sql)
            queryset = query.fetchall()
            retCols = query.keys()[:-1]
            row_data = []
            for item in queryset:
                tmp = {}
                for index in range(len(retCols)):
                    tmp[retCols[index]] = item[index]
                row_data.append(tmp)
            yaml_str = yaml_dict.get(request.statusType, None)
            if not yaml_str:
                return
            yaml_str = StatsRepository.get_branch_timeliness_yaml(yaml_str, self.context.app.config.get(
                Config.QC_STATS_BRANCH_TARGET_FIELD) or "")
            cfg = BIFormConfig.fromYaml(yaml_str)
            processor = BIDataProcess(cfg, row_data)

            file_name = "{}{}".format("全院病案指标", datetime.now().strftime("%Y-%m-%d")) + ".xlsx"
            file_id = self.get_file_id(file_name)
            path_file_name = os.path.join(export_path, file_id + ".xlsx")
            processor.toExcel(path=path_file_name, sheet_name="全院病案指标")
            response["id"] = file_id
            response["fileName"] = file_name
            return response


class GetBranchTimelinessRateDetail(MyResource):

    @pre_request(request, GetBranchTimelinessRateDetailReq)
    def post(self):
        """
        全院病案指标-指标明细
        :param request:
        :param context:
        :return:
        """
        response = {"headers": [], "data": [], "pageInfo": {}}
        target_field = BRANCH_TARGET_FIELD_DICT.get(request.targetName)
        if not target_field:
            self.logger.error("request.targetName: %s is error.", request.targetName)
            return get_error_resp("参数错误")

        with self.context.app.mysqlConnection.session() as session:
            to_web_data, total = StatsRepository.GetBranchTimelinessRateData(session, request, target_field)
        yaml_conf = BRANCH_TIMELINESS_DETAIL_LIST_YAML
        if request.targetName in ("入院记录24小时内完成率", "手术记录24小时内完成率", "出院记录24小时内完成率", "病案首页24小时内完成率"):
            yaml_conf = BRANCH_TIMELINESS_DETAIL_TIME_LIST_YAML
        cfg = BIFormConfig.fromYaml(yaml_conf)
        processor = BIDataProcess(cfg, to_web_data)
        header, result = processor.toWeb()
        QCUtil.format_header_data(response, header, result)

        response["pageInfo"]["start"] = request.start or 0
        response["pageInfo"]["size"] = request.size or 10
        response["pageInfo"]["total"] = total
        return response


class GetBranchTimelinessRateDetailFormula(MyResource):

    @pre_request(request, GetBranchTimelinessRateDetailReq)
    def post(self):
        """
        全院病案指标-指标明细-公式数据
        :param request:
        :param context:
        :return:
        """
        target_dict = {
            1: '全院病案指标',
            2: '科室病案指标',
            3: '医生病案指标',
            4: '病区病案指标'
        }
        dimension_dict = {
            1: 'branch',
            2: 'department',
            3: 'attend',
            4: 'ward'
        }
        response = {}
        with self.context.app.mysqlConnection.session() as session:
            dept_or_ward = request.department or request.ward or "全部"
            if dept_or_ward == "总计":
                dept_or_ward = "全部"
            args = [target_dict.get(request.statusType), request.startTime or "2000-01-01", request.endTime or "2030-12-31",
                    request.branch if request.branch and request.branch != "全院" else "全部",
                    dept_or_ward, request.attend if request.attend and request.attend != "总计" else "全部"]
            call_proc_sql = """call pr_case_medical_index('%s','%s', '%s', '%s', '%s', '%s')""" % tuple(args)
            query = session.execute(call_proc_sql)
            queryset = query.fetchall()
            retCols = query.keys()[:-1]
            find_col = False
            col_current = 3 if not request.attend else 4
            for col in range(col_current, len(retCols), 3):
                if retCols[col] == request.targetName:
                    col_current = col
                    find_col = True
            if not find_col:
                return
            dim = dimension_dict.get(request.statusType)
            dim_value = getattr(request, dim, '总计')
            select_index = 0 if dim != 'attend' else 1
            for item in queryset:
                if item[select_index] == dim_value:
                    response["moleculeNum"] = str(item[col_current-2] or 0)
                    response["denominatorNum"] = str(item[col_current-1] or 0)
                    break

            query_formula_sql = '''select mumerator, denominator from dim_firstpageindex where name = "%s"''' % request.targetName
            query = session.execute(query_formula_sql)
            queryset = query.fetchone()
            if queryset:
                response["molecule"] = queryset[0]
                response["denominator"] = queryset[1]

        return response
        

class GetBranchTimelinessRateDetailExport(MyResource):

    @pre_request(request, GetBranchTimelinessRateDetailReq)
    def post(self):
        """
        全院病案指标-指标明细-导出
        :param request:
        :param context:
        :return:
        """
        response = {}
        target_field = BRANCH_TARGET_FIELD_DICT.get(request.targetName)
        if not target_field:
            self.logger.error("request.targetName: %s is error.", request.targetName)
            return get_error_resp("参数错误")

        with self.context.app.mysqlConnection.session() as session:
            to_excel_data, total = StatsRepository.GetBranchTimelinessRateData(session, request, target_field, is_export=1)
        yaml_conf = BRANCH_TIMELINESS_DETAIL_LIST_YAML
        if request.targetName in ("入院记录24小时内完成率", "手术记录24小时内完成率", "出院记录24小时内完成率", "病案首页24小时内完成率"):
            yaml_conf = BRANCH_TIMELINESS_DETAIL_TIME_LIST_YAML
        cfg = BIFormConfig.fromYaml(yaml_conf)
        processor = BIDataProcess(cfg, to_excel_data)

        file_name = "【{}】{}{}".format(request.targetName, "指标明细", datetime.now().strftime("%Y-%m-%d")) + ".xlsx"
        file_id = self.get_file_id(file_name)
        path_file_name = os.path.join(export_path, file_id + ".xlsx")
        processor.toExcel(path=path_file_name, sheet_name=request.targetName)
        response["id"] = file_id
        response["fileName"] = file_name
        return response


class GetDoctorArchivingRate(MyResource):

    @pre_request(request, GetBranchTimelinessRateDetailReq)
    def post(self):
        """
        统计医生归档率
        :param request:
        :param context:
        :return:
        """
        response = {"data": []}
        sizeCount = 0
        start, size = QCUtil.getRequestStartAndSize(request)
        current = 0
        # startDate = request.startTime.ToDatetime()
        branch = request.branch or "全部"
        department = request.department or "全部"
        doctor = request.doctor or "全部"
        doctorFR = request.doctorFR or "全部"
        args = [request.startTime, request.endTime, 'doctor', branch, department, doctor, doctorFR]
        sortKey = SortDict[request.sortKey] if request.sortKey else 19

        with self.context.app.mysqlConnection.session() as cursor:
            call_proc_sql = """call pr_case_archiverate('%s', '%s', '%s', '%s', '%s', '%s', '%s')""" % tuple(args)
            query = cursor.execute(call_proc_sql)
            self.logger.info("GetDoctorArchivingRate, call_proc_sql: %s", call_proc_sql)
            queryset = query.fetchall()
            totalCount = sum([1 for item in queryset if not item[21]])
            if queryset:
                # sort result:
                # 过滤掉F/R列
                if not request.enableFR:
                    filter(lambda x: x[20] == "--", queryset)
                sortedResult = QCUtil.sortArchivingRateResult(queryset, sortKey, request.sortWay)
                tmp_list = []
                for item in sortedResult:
                    if (item[0], item[1]) in tmp_list:
                        continue
                    tmp_list.append((item[0], item[1]))
                    current += 1
                    if current <= start:
                        continue
                    # 控制分页
                    if current > start + size:
                        break
                    else:
                        sizeCount += 1
                        protoItem = {}  # response.data.add()
                        protoItem["department"] = item[0] if item[0] else ""
                        protoItem["doctor"] = item[1] if item[1] else ""
                        protoItem["dischargeCount"] = int(item[2]) if item[2] else 0
                        protoItem["applyCount"] = int(item[3]) if item[3] else 0
                        protoItem["refusedCount"] = int(item[4]) if item[4] else 0
                        protoItem["archivedCount"] = int(item[5]) if item[5] else 0
                        protoItem["unreviewCount"] = int(item[6]) if item[6] else 0
                        protoItem["archivingRate"] = float(item[7]) if item[7] else 0
                        protoItem["refusedRate"] = float(item[8]) if item[8] else 0
                        protoItem["finishRate"] = float(item[9]) if item[9] else 0
                        protoItem["imperfect"] = float(item[10]) if item[10] else 0

                        protoItem["secondApply"] = int(item[11]) if item[11] else 0
                        protoItem["secondNoApply"] = int(item[12]) if item[12] else 0
                        protoItem["secondArchivingRate"] = float(item[13]) if item[13] else 0
                        protoItem["thirdApply"] = int(item[14]) if item[14] else 0
                        protoItem["thirdNoApply"] = int(item[15]) if item[15] else 0
                        protoItem["thirdArchivingRate"] = float(item[16]) if item[16] else 0
                        protoItem["seventhApply"] = int(item[17]) if item[17] else 0
                        protoItem["seventhNoApply"] = int(item[18]) if item[18] else 0
                        protoItem["seventhArchivingRate"] = float(item[19]) if item[19] else 0
                        protoItem["doctorFR"] = item[20] if item[20] else ""
                        protoItem["doctorFRFlag"] = item[21] if item[21] else ""

                        protoItem["timelyRate"] = str(item[22] if item[22] else 0)
                        protoItem["fixRate"] = str(item[23] if item[23] else 0)
                        response["data"].append(protoItem)

        response["total"] = totalCount
        response["size"] = sizeCount
        response["start"] = start
        return response


class ExportDoctorArchivingRate(MyResource):

    @pre_request(request, GetBranchTimelinessRateDetailReq)
    def post(self):
        """
        导出医生归档率
        :param request:
        :param context:
        :return:
        """
        response = {}
        branch = request.branch or "全部"
        department = request.department or "全部"
        doctor = request.doctor or "全部"
        doctorFR = request.doctorFR or "全部"
        args = [request.startTime, request.endTime, 'doctor', branch, department, doctor, doctorFR]
        sortKey = SortDict[request.sortKey] if request.sortKey else 19

        excelTitle = "{0}至{1}各ATTENDING组出院病人病历完成情况".format(request.startTime, request.endTime)
        exportName = "医生归档率{0}至{1}.xlsx".format(request.startTime, request.endTime)
        with self.context.app.mysqlConnection.session() as cursor:
            fileId = QCUtil.getFileId()
            filename = "%s.xlsx" % fileId
            fullname = os.path.join(export_path, filename)
            wb = Workbook()
            call_proc_sql = """call pr_case_archiverate('%s', '%s', '%s', '%s', '%s', '%s', '%s')""" % tuple(args)
            query = cursor.execute(call_proc_sql)
            self.logger.info("ExportDoctorArchivingRate, call_proc_sql: %s", call_proc_sql)
            queryset = query.fetchall()
            retCols = query.keys()
            if queryset:
                if not request.enableFR:
                    filter(lambda item: item[20] == "--", queryset)
                sortedResult = QCUtil.sortArchivingRateResult(queryset, sortKey, request.sortWay)
                QCUtil.generateWorkSheet(sortedResult, wb, retCols, title=excelTitle, enableFR=request.enableFR)
            wb.save(fullname)

        response["id"] = str(fileId)
        response["fileName"] = exportName
        return response
    
    
class GetDirectorArchivingRate(MyResource):

    @pre_request(request, GetBranchTimelinessRateDetailReq)
    def post(self):
        """
        获取科主任统计接口
        :param request:
        :param context:
        :return:
        """
        response = {"data": []}
        sizeCount = 0
        start, size = QCUtil.getRequestStartAndSize(request)
        current = 0
        branch = request.branch or "全部"
        department = request.department or "全部"
        args = [request.startTime, request.endTime, branch, department]
        # sortKey = SortDict[request.sortKey] if request.sortKey else 7
        # 只根据序号Id来排序
        sortKey = 7
        with self.context.app.mysqlConnection.session() as cursor:
            call_proc_sql = """call pr_case_aindex('%s', '%s', '%s', '%s')""" % tuple(args)
            query = cursor.execute(call_proc_sql)
            self.logger.info("GetDirectorArchivingRate, call_proc_sql: %s", call_proc_sql)
            queryset = query.fetchall()
            totalCount = len(queryset)
            if queryset:
                # sort result:
                sortedResult = QCUtil.sortDirectorStatsResult(queryset, sortKey, request.sortWay)
                for item in sortedResult:
                    current += 1
                    if current <= start:
                        continue
                    # 控制分页
                    if current > start + size:
                        break
                    else:
                        sizeCount += 1
                        protoItem = {}  # response.data.add()
                        protoItem["department"] = item[0] if item[0] else ""
                        protoItem["doctor"] = item[1] if item[1] else ""
                        protoItem["dischargeCount"] = int(item[2]) if item[2] else 0
                        protoItem["primaryDiagValidRate"] = float(item[3]) if item[3] else 0
                        protoItem["minorDiagValidRate"] = float(item[4]) if item[4] else 0
                        protoItem["primaryOperValidRate"] = float(item[5]) if item[5] else 0
                        protoItem["minorOperValidRate"] = float(item[6]) if item[6] else 0
                        response["data"].append(protoItem)

        response["total"] = totalCount
        response["size"] = sizeCount
        response["start"] = start
        return response


class GetDoctorArchivingRateCase(MyResource):

    @pre_request(request, GetDoctorArchivingRateCaseReq)
    def post(self):
        """
        获取医生归档率详情病历
        :param request:
        :param context:
        :return:
        """
        response = {"items": []}
        sizeCount = 0
        queryset, totalCount, start = QCUtil.query_docror_archiving_data(self.app, request)
        if not queryset:
            return response
        for item in queryset:
            sizeCount += 1
            protoItem = {}  # response.items.add()
            unmarshalDoctorArchivingRateCase(protoItem, item)
            response["items"].append(protoItem)

        response["total"] = totalCount
        response["size"] = sizeCount
        response["start"] = start
        return response


class ExportDoctorArchivingRateCase(MyResource):

    @pre_request(request, GetDoctorArchivingRateCaseReq)
    def post(self):
        """
        导出医生归档率详情病历
        :param request:
        :param context:
        :return:
        """
        response = {}
        queryset, totalCount, start = QCUtil.query_docror_archiving_data(self.app, request)
        fileId = QCUtil.getFileId()
        filename = "%s.xlsx" % fileId
        fullname = os.path.join(export_path, filename)
        exportName = "医生归档率明细{0}至{1}.xlsx".format(request.startTime, request.endTime)
        wb = Workbook()
        patient_id_name = self.context.app.config.get(Config.QC_PATIENT_ID_NAME)
        titles = [patient_id_name, "姓名", "入院日期", "出院日期", "出院科室", "医生", "住院天数",
                  "申请标记", "首次申请日期", "首次申请人", "审核标记", "审核日期", "审核人", "审核说明", "是否存在当前出院科室的问题", "死亡标记",
                  "是否符合2日归档", "是否符合3日归档", "是否符合7日归档"]
        ws = wb.active
        ws.append(tuple(titles))
        for item in queryset:
            rowData = []
            rowData.append(item[2] if item[2] else "")
            rowData.append(item[3] if item[3] else "")
            rowData.append(item[9].strftime("%Y-%m-%d") if item[9] else "")
            rowData.append(item[10].strftime("%Y-%m-%d") if item[10] else "")
            rowData.append(item[7] if item[7] else "")
            rowData.append(item[8] if item[8] else "")
            rowData.append(item[12] if item[12] else "0")
            rowData.append(item[13] if item[13] else "")
            rowData.append(item[14].strftime("%Y-%m-%d") if item[14] else "")
            rowData.append(item[15] if item[15] else "")
            rowData.append(item[16] if item[16] else "")
            rowData.append(item[18].strftime("%Y-%m-%d") if item[18] else "")
            rowData.append(item[17] if item[17] else "")

            hasDischargeDeptProblemFlag = "是" if item[23] else "否"
            if item[16] == "已退回":
                # 只有已退回状态的才显示审核说明
                # rowData.append(item[28] if item[28] else "")
                rowData.append("")
                rowData.append(hasDischargeDeptProblemFlag)
            else:
                rowData.append("")
                rowData.append("")

            deadFlag = "死亡" if item[19] and item[19] == 1 else "未亡"
            rowData.append(deadFlag)
            rowData.append(item[20] if item[20] else "")
            rowData.append(item[21] if item[21] else "")
            rowData.append(item[22] if item[22] else "")

            ws.append(tuple(rowData))
        wb.save(fullname)
        response["id"] = str(fileId)
        response["fileName"] = exportName
        return response


class ExportDirectorArchivingRateCase(MyResource):

    @pre_request(request, GetDoctorArchivingRateCaseReq)
    def post(self):
        """
        导出科主任统计病历详情病历
        :return:
        """
        response = {}
        queryset, totalCount, start = QCUtil.query_docror_archiving_data(self.app, request)

        fileId = QCUtil.getFileId()
        filename = "%s.xlsx" % fileId
        fullname = os.path.join(export_path, filename)
        exportName = "{0}至{1}科主任指标统计明细.xlsx".format(request.startTime, request.endTime)

        wb = Workbook()
        patient_id_name = self.context.app.config.get(Config.QC_PATIENT_ID_NAME)
        titles = [patient_id_name, "姓名", "入院日期", "出院日期", "出院科室", "医生", "住院天数",
                  "首页主诊断是否正确", "首页次诊断是否完整", "首页主手术/操作填写是否准确", "首页次手术/操作填写是否完整"]
        ws = wb.active
        ws.append(tuple(titles))

        for item in queryset:
            rowData = []
            rowData.append(item[2] if item[2] else "")
            rowData.append(item[3] if item[3] else "")
            rowData.append(item[9].strftime("%Y-%m-%d") if item[9] else "")
            rowData.append(item[10].strftime("%Y-%m-%d") if item[10] else "")
            rowData.append(item[7] if item[7] else "")
            rowData.append(item[8] if item[8] else "")
            rowData.append(item[12] if item[12] else "0")
            rowData.append(item[24] if item[24] else "")
            rowData.append(item[25] if item[25] else "")
            rowData.append(item[26] if item[26] else "")
            rowData.append(item[27] if item[27] else "")
            ws.append(tuple(rowData))
        wb.save(fullname)
        response["id"] = str(fileId)
        response["fileName"] = exportName
        return response


class ExportDirectorArchivingRate(MyResource):

    @pre_request(request, GetDoctorArchivingRateCaseReq)
    def post(self):
        """
        导出科主任统计接口
        :param request:
        :param context:
        :return:
        """
        response = {}
        branch = request.branch or "全部"
        department = request.department or "全部"
        args = [request.startTime, request.endTime, branch, department]
        sortKey = 7

        excelTitle = "{0}至{1}科主任指标统计".format(request.startTime, request.endTime)
        exportName = "{0}至{1}科主任指标统计.xlsx".format(request.startTime, request.endTime)
        with self.context.app.mysqlConnection.session() as cursor:
            fileId = QCUtil.getFileId()
            filename = "%s.xlsx" % fileId
            fullname = os.path.join(export_path, filename)
            wb = Workbook()
            call_proc_sql = """call pr_case_aindex('%s', '%s', '%s', '%s')""" % tuple(args)
            query = cursor.execute(call_proc_sql)
            queryset = query.fetchall()
            retCols = query.keys()
            if queryset:
                totalCount = len(queryset)
                self.logger.info("totalCount is : %s" % totalCount)
                sortedResult = QCUtil.sortDirectorStatsResult(queryset, sortKey, request.sortWay)
                QCUtil.generateDirectorWorkSheet(sortedResult, wb, retCols, title=excelTitle)
            wb.save(fullname)
        response["id"] = str(fileId)
        response["fileName"] = exportName
        return response


class GetMedicalIndicatorStats(MyResource):

    @pre_request(request, GetDoctorArchivingRateCaseReq)
    def post(self):
        """
        获取病案指标统计接口
        :param request:
        :param context:
        :return:
        """
        response = {"data": []}
        sizeCount = 0
        start, size = QCUtil.getRequestStartAndSize(request)
        totalCount = 0
        current = 0
        removeAllFlag = True
        # startDate = request.startTime.ToDatetime()
        branch = request.branch or "全部"
        department = request.department or "全部"
        doctor = request.doctor or "全部"
        args = [request.startTime, request.endTime, branch, department, doctor]
        if branch == "全部" and department == "全部" and doctor == "全部":
            removeAllFlag = False

        sortKey = 14
        self.logger.info(json.dumps(args, ensure_ascii=False))
        with self.context.app.mysqlConnection.session() as cursor:
            call_proc_sql = """call pr_case_medical_index('%s', '%s', '%s', '%s', '%s')""" % tuple(args)
            query = cursor.execute(call_proc_sql)
            queryset = query.fetchall()
            if queryset:
                sortedResult = QCUtil.sortMedicalIndicatorResult(queryset, sortKey, request.sortWay, removeAllFlag)
                totalCount = len(sortedResult)
                self.logger.info("sortedResult len is : %s" % totalCount)
                for item in sortedResult:
                    current += 1
                    if current <= start:
                        continue
                    # 控制分页
                    if current > start + size:
                        break
                    else:
                        sizeCount += 1
                        protoItem = {}  # response.data.add()
                        protoItem["branch"] = item[0] if item[0] else ""
                        protoItem["department"] = item[1] if item[1] else ""
                        protoItem["doctor"] = item[2] if item[2] else ""
                        protoItem["admitRecord24HRate"] = item[3] if item[3] else ""
                        protoItem["operationRecord24HRate"] = item[4] if item[4] else ""
                        protoItem["dischargeRecord24HRate"] = item[5] if item[5] else ""
                        protoItem["firstPage24HRate"] = item[6] if item[6] else ""
                        protoItem["operationRecordFinishRate"] = item[7] if item[7] else ""
                        protoItem["roundRecordFinishRate"] = item[8] if item[8] else ""
                        protoItem["rescueRecordFinishRate"] = item[9] if item[9] else ""
                        protoItem["secondArchivingRate"] = item[10] if item[10] else ""
                        protoItem["archivingFinishRate"] = item[11] if item[11] else ""
                        protoItem["unreasonableCaseRate"] = item[12] if item[12] else ""
                        protoItem["consentSignedRate"] = item[13] if item[13] else ""
                        response["data"].append(protoItem)

        response["total"] = totalCount
        response["size"] = sizeCount
        response["start"] = start
        return response


class ExportMedicalIndicatorStats(MyResource):

    @pre_request(request, GetDoctorArchivingRateCaseReq)
    def post(self): 
        """
        导出病案指标统计接口
        :param request:
        :param context:
        :return:
        """
        response = {}
        removeAllFlag = True
        branch = request.branch or "全部"
        department = request.department or "全部"
        doctor = request.doctor or "全部"
        args = [request.startTime, request.endTime, branch, department, doctor]
        if branch == "全部" and department == "全部" and doctor == "全部":
            removeAllFlag = False
        sortKey = 14

        excelTitle = "{0}至{1}病案指标统计".format(request.startTime, request.endTime)
        exportName = "{0}至{1}病案指标统计.xlsx".format(request.startTime, request.endTime)
        with self.context.app.mysqlConnection.session() as cursor:
            fileId = QCUtil.getFileId()
            filename = "%s.xlsx" % fileId
            fullname = os.path.join(export_path, filename)
            wb = Workbook()
            call_proc_sql = """call pr_case_medical_index('%s', '%s', '%s', '%s', '%s')""" % tuple(args)
            query = cursor.execute(call_proc_sql)
            queryset = query.fetchall()
            retCols = query.keys()
            if queryset:
                sortedResult = QCUtil.sortMedicalIndicatorResult(queryset, sortKey, request.sortWay, removeAllFlag)
                QCUtil.generateMedicalIndicatorWorkSheet(sortedResult, wb, retCols, title=excelTitle)
            wb.save(fullname)
        response["id"] = str(fileId)
        response["fileName"] = exportName
        return response


class GetMedicalIndicatorStatsCase(MyResource):

    @pre_request(request, GetDoctorArchivingRateCaseReq)
    def post(self):
        """
        获取病案指标详情接口
        :param request:
        :param context:
        :return:
        """
        response = {"data": []}
        sizeCount = 0
        queryset, totalCount, start = QCUtil.query_medical_indicator_data(request)
        if not queryset:
            return response
        for item in queryset:
            sizeCount += 1
            protoItem = {}  # response.data.add()
            protoItem["caseId"] = item[1] if item[1] else ""
            protoItem["patientId"] = item[2] if item[2] else ""
            protoItem["name"] = item[3] if item[3] else ""
            protoItem["caseStatus"] = self.parseCaseStatus(item[12]) if item[12] else ""
            protoItem["department"] = item[7] if item[7] else ""
            protoItem["doctor"] = item[8] if item[8] else ""
            protoItem["admitTime"] = item[9].strftime("%Y-%m-%d") if item[9] else ""
            protoItem["dischargeTime"] = item[10].strftime("%Y-%m-%d") if item[10] else ""

            protoItem["admitRecord24H"] = item[13] if item[13] else ""
            protoItem["operationRecord24H"] = item[14] if item[14] else ""
            protoItem["dischargeRecord24H"] = item[15] if item[15] else ""
            protoItem["firstPage24H"] = item[16] if item[16] else ""
            protoItem["operationRecordFinish"] = item[17] if item[17] else ""
            protoItem["roundRecordFinish"] = item[18] if item[18] else ""
            protoItem["rescueRecordFinish"] = item[19] if item[19] else ""
            protoItem["secondArchiving"] = item[20] if item[20] else ""
            protoItem["archivingFinish"] = item[21] if item[21] else ""
            protoItem["unreasonableCase"] = item[22] if item[22] else ""
            protoItem["consentSigned"] = item[23] if item[23] else ""
            response["data"].append(protoItem)

        response["total"] = totalCount
        response["size"] = sizeCount
        response["start"] = start
        return response


class ExportMedicalIndicatorStatsCase(MyResource):

    @pre_request(request, GetDoctorArchivingRateCaseReq)
    def post(self):
        """
        导出医生归档率详情病历
        :param request:
        :param context:
        :return:
        """
        response = {}

        queryset, totalCount, start = QCUtil.query_medical_indicator_data(request)
        fileId = QCUtil.getFileId()
        filename = "%s.xlsx" % fileId
        fullname = os.path.join(export_path, filename)
        exportName = "{0}至{1}指标统计明细.xlsx".format(request.startTime, request.endTime)
        wb = Workbook()
        patient_id_name = self.context.app.config.get(Config.QC_PATIENT_ID_NAME)
        titles = [patient_id_name, "姓名", "入院日期", "出院日期", "病历状态", "科室", "医生", "入院记录24小时完成",
                  "手术记录24小时完成", "出院记录24小时完成", "病案首页24小时完成", "手术相关记录完整",
                  "医生查房记录完整", "患者抢救记录及时完成", "出院患者病历2日归档", "出院患者病历归档完整",
                  "不合理复制病历发生", "知情同意书规范签署"]
        ws = wb.active
        ws.append(tuple(titles))

        totalCount = len(queryset)
        self.logger.info("queryset len is : %s" % totalCount)
        if queryset:
            for item in queryset:
                rowData = []
                rowData.append(item[2] if item[2] else "")
                rowData.append(item[3] if item[3] else "")
                rowData.append(item[9].strftime("%Y-%m-%d") if item[9] else "")
                rowData.append(item[10].strftime("%Y-%m-%d") if item[10] else "")
                caseStatus = self.parseCaseStatus(item[12]) if item[12] else ""
                rowData.append(caseStatus)
                rowData.append(item[7] if item[7] else "")
                rowData.append(item[8] if item[8] else "")

                for idx in range(13, 24):
                    rowData.append(item[idx] if item[idx] else "")

                ws.append(tuple(rowData))
        wb.save(fullname)
        response["id"] = str(fileId)
        response["fileName"] = exportName
        return response


class GetStatsCaseTag(MyResource):

    @pre_request(request, ["input"])
    def get(self):
        """
        获取病历标签字典
        :param request:
        :param context:
        :return:
        """
        response = {"data": []}
        query_name = request.input or ""

        with self.context.app.mysqlConnection.session() as cursor:
            query_tag_sql = '''select * from tags_qc'''
            if query_name:
                query_tag_sql += ''' where name like "%{}%"'''.format(query_name)
            query_tag_sql += ''' order by no'''
            self.logger.info("GetStatsCaseTag, execute query_tag_sql: %s", query_tag_sql)
            query = cursor.execute(query_tag_sql)
            data = query.fetchall()

            for item in data:
                protoItem = {}  # response.data.add()
                protoItem["id"] = item[0]
                protoItem["name"] = item[1] or ""
                protoItem["code"] = item[2] or ""
                protoItem["status"] = int(item[3])
                protoItem["orderNo"] = int(item[4])

        return response


class GetFirstPageProblems(MyResource):

    @pre_request(request, GetDoctorArchivingRateCaseReq)
    def post(self):
        """ 病案首页问题统计
        """
        response = {"data": []}
        start, size = QCUtil.getRequestPageAndCount(request)
        sortKey = {
            "totalCount": "firstProblemCount",
            "requiredCount": "requiredProblemCount",
            "optionalCount": "optionalProblemCount"
        }
        if not request.sortKey:
            request.sortKey = "totalCount"
        with self.context.app.mysqlConnection.session() as cursor:
            query_sql = '''select c.caseId, c.patientId, c.name, c.outDeptName, c.wardName, c.attendDoctor, c.admitTime, 
            c.dischargeTime, c.status, ar.firstProblemCount, ar.requiredProblemCount, ar.optionalProblemCount 
            from `case` c left outer join audit_record ar on c.audit_id = ar.id 
            left outer join dim_dept_statis dds on dds.deptid = c.outDeptId '''
            query_count_sql = '''select count(*) from `case` c left outer join audit_record ar on c.audit_id = ar.id 
            left outer join dim_dept_statis dds on dds.deptid = c.outDeptId '''
            filter_sql = "where 1 = 1"
            if request.startTime:
                filter_sql += " and c.dischargeTime >= '%s'" % request.startTime
            if request.endTime:
                filter_sql += " and c.dischargeTime <= '%s 23:59:59'" % request.endTime
            if request.outDept:
                filter_sql += " and dds.statis_name = '%s'" % request.outDept
            if request.status:
                filter_sql += " and c.status = '%s'" % request.status
            if request.problemFlag:
                if request.problemFlag == 1:
                    filter_sql += " and ar.firstProblemCount > 0"
                elif request.problemFlag == 2:
                    filter_sql += " and ar.firstProblemCount = 0"
            if request.input:
                filter_sql += """ and c.patientId like '%{name}%' or c.name like '%{name}%' 
                or c.attendDoctor like '%{name}%'""".format(name=request.input)
            order_by_sql = ""
            if request.sortKey:
                if not request.sortWay or request.sortWay.lower() not in ["asc", "desc"]:
                    request.sortWay = "DESC"
                if request.sortKey in ["patientId", "dischargeTime"]:
                    order_by_sql += " order by c.%s %s" % (request.sortKey, request.sortWay)
                elif sortKey.get(request.sortKey):
                    order_by_sql += " order by ar.%s %s" % (sortKey.get(request.sortKey), request.sortWay)

            query_sql += filter_sql
            query_count_sql += filter_sql
            query_sql += order_by_sql
            query_sql += " limit %s, %s" % (start, size)
            query_count = cursor.execute(query_count_sql)
            response["total"] = query_count.fetchone()[0]
            query = cursor.execute(query_sql)
            data = query.fetchall()

            for item in data:
                protoItem = {}  # response.data.add()
                protoItem["caseId"] = item[0] or ""
                protoItem["patientId"] = item[1] or ""
                protoItem["name"] = item[2] or ""
                protoItem["outDept"] = item[3] or ""
                protoItem["ward"] = item[4] or ""
                protoItem["attend"] = item[5] or ""
                protoItem["admitTime"] = item[6].strftime("%Y-%m-%d") if item[6] else ""
                protoItem["dischargeTime"] = item[7].strftime("%Y-%m-%d") if item[7] else ""
                protoItem["status"] = item[8] or 0
                protoItem["totalCount"] = item[9] or 0
                protoItem["requiredCount"] = item[10] or 0
                protoItem["optionalCount"] = item[11] or 0
                response["data"].append(protoItem)
        return response


class GetFirstPageScoreStats(MyResource):

    @pre_request(request, StartEndTimeReq)
    def post(self):
        """
        各科室首页分数统计接口
        :param request:
        :param context:
        :return:
        """
        response = {"data": []}

        args = ["各科室首页分数统计", request.startTime, request.endTime, ""]
        with self.context.app.mysqlConnection.session() as cursor:
            call_proc_sql = '''call pr_case_firstpagescore("%s", "%s", "%s", "%s")''' % tuple(args)
            query = cursor.execute(call_proc_sql)
            self.logger.info("GetFirstPageScoreStats, call_proc_sql: %s", call_proc_sql)
            queryset = query.fetchall()
            if queryset:
                self.logger.info("GetFirstPageScoreStats len queryset: %s", len(queryset))
                for item in queryset:
                    protoItem = {}  # response.data.add()
                    protoItem["department"] = item[0] if item[0] else ""
                    protoItem["maxScore"] = float(item[1]) if item[1] else 0
                    protoItem["minScore"] = float(item[2]) if item[2] else 0
                    protoItem["avgScore"] = float(item[3]) if item[3] else 0
                    protoItem["hospitalAvgScore"] = float(item[4]) if item[4] else 0
                    response["data"].append(protoItem)

        return response


class GetFirstPageScoreDistribution(MyResource):

    @pre_request(request, StartEndTimeReq)
    def post(self):
        """
        首页分数分布统计接口
        :param request:
        :param context:
        :return:
        """
        response = {"data": []}

        args = ["首页分数分布", request.startTime, request.endTime, ""]
        with self.context.app.mysqlConnection.session() as cursor:
            call_proc_sql = '''call pr_case_firstpagescore("%s", "%s", "%s", "%s")''' % tuple(args)
            query = cursor.execute(call_proc_sql)
            self.logger.info("GetFirstPageScoreDistribution, call_proc_sql: %s", call_proc_sql)
            queryset = query.fetchall()
            if queryset:
                self.logger.info("GetFirstPageScoreDistribution len queryset: %s", len(queryset))
                for item in queryset:
                    protoItem = {}  # response.data.add()
                    protoItem["area"] = item[1] if item[1] else ""
                    protoItem["caseCount"] = int(item[0]) if item[0] else 0
                    protoItem["areaIndex"] = int(item[2]) if item[2] else 0
                    response["data"].append(protoItem)

        return response


class GetFirstPageIndicateStats(MyResource):

    @pre_request(request, GetFirstPageIndicateStatsReq)
    def post(self):
        """
        首页通用指标统计接口
        :param request:
        :param context:
        :return:
        """
        response = {"items": []}
        if not request.indicateName:
            self.logger.error("GetFirstPageIndicateStats: param indicateName is required.")
            return get_error_resp("GetFirstPageIndicateStats: param indicateName is required.")

        limitSize = request.count if request.count else 10
        params = request.indicateParams if request.indicateParams else ""
        args = [request.indicateName, request.startTime, request.endTime, params]
        with self.context.app.mysqlConnection.session() as cursor:
            call_proc_sql = '''call pr_case_firstpagescore("%s", "%s", "%s", "%s")''' % tuple(args)
            query = cursor.execute(call_proc_sql)
            self.logger.info("GetFirstPageIndicateStats, call_proc_sql: %s", call_proc_sql)
            queryset = query.fetchall()
            if queryset:
                self.logger.info("GetFirstPageIndicateStats len queryset: %s", len(queryset))
                size = 0
                for item in queryset:
                    size += 1
                    if size > limitSize:
                        break
                    protoItem = {}  # response.items.add()
                    protoItem["name"] = item[0] if item[0] else ""
                    protoItem["count"] = int(item[1]) if item[1] else 0
                    response["items"].append(protoItem)

        return response


class GetFirstPageProblemConfig(MyResource):

    def get(self):
        """
        获取首页问题配置接口
        :param request:
        :param context:
        :return:
        """
        response = {"data": []}

        with self.context.app.mysqlConnection.session() as cursor:
            query_sql = "select distinct name ,status, sort from dim_firstpagescore where is_select = 1 order by sort;"
            query = cursor.execute(query_sql)
            self.logger.info("GetFirstPageProblemConfig, query_sql: %s", query_sql)
            ret = query.fetchall()
            for x in ret:
                protoItem = {}  # response.data.add()
                protoItem["name"] = x[0] if x[0] else ""
                protoItem["status"] = int(x[1]) if x[1] else 0
                protoItem["sortIndex"] = int(x[2]) if x[2] else 0
                response["data"].append(protoItem)

        return response


class ModifyFirstPageProblemConfig(MyResource):

    @pre_request(request, ["data:list"])
    def post(self):
        """
        修改首页问题排序接口
        :param request:
        :param context:
        :return:
        """
        if request.data:
            with self.context.app.mysqlConnection.session() as cursor:
                for item in request.data:
                    updateSql = "update dim_firstpagescore set status = {}, sort = {} where name = '{}';".format(
                        item["status"], item["sortIndex"], item["name"])
                    self.logger.info(updateSql)
                    cursor.execute(updateSql)

        return g.result


class GetBatchFirstPageIndicateStats(MyResource):

    @pre_request(request, StartEndTimeReq)
    def post(self):
        """
        首页通用指标统计接口
        :param request:
        :param context:
        :return:
        """
        response = {"items": []}
        if not request.indicates:
            return get_error_resp("indicates can not be empty.")

        resultDict = {}
        plist = []
        start = time.time()
        indicatorList = []
        # self.logger.error("GetBatchFirstPageIndicateStats, start time: %s, len: %s", start, len(request.indicates))
        for item in request.indicates:
            params = item.get("params", "")
            indicatorID = "%s-%s" % (item["name"], params)
            indicatorList.append(indicatorID)
            p = threading.Thread(target=self.processStats,
                                 args=(indicatorID, item, request.startTime, request.endTime, resultDict))
            p.start()
            plist.append(p)

        for x in plist:
            x.join(90)
            # x.close()

        self.logger.error("GetBatchFirstPageIndicateStats end time: %s, use time: %s", time.time(), time.time() - start)
        # get the result data
        for key in indicatorList:
            result = resultDict[key]
            protoItem = {"items": []}  # response.items.add()
            keyArr = key.split("-")
            protoItem["name"] = keyArr[0]
            if len(keyArr) == 2 and keyArr[1]:
                protoItem["params"] = keyArr[1]
            for retItem in result:
                protoItemVal = {}  # protoItem.items.add()
                protoItemVal["name"] = retItem.get("dept") if retItem.get("dept") else ""
                protoItemVal["count"] = retItem.get("count") if retItem.get("count") else 0
                protoItem["items"].append(protoItemVal)
            response["items"].append(protoItem)

        return response

    def processStats(self, indicatorID, indicatorItem, startTime, endTime, resultDict):
        """
        process Stats
        # 并行处理单元
        :return:
        """
        result = self.mysqlStats(indicatorItem, startTime, endTime)
        resultDict[indicatorID] = result

    def mysqlStats(self, indicatorItem, startTime, endTime):
        """
        mysql Stats
        :return:
        """
        result = []
        limitSize = indicatorItem.get("count", 10)
        params = indicatorItem.get("params", "")
        args = [indicatorItem["name"], startTime, endTime, params]
        with self.context.app.mysqlConnection.session() as cursor:
            call_proc_sql = '''call pr_case_firstpagescore("%s", "%s", "%s", "%s")''' % tuple(args)
            query = cursor.execute(call_proc_sql)
            self.logger.info("mysqlStats, call_proc_sql: %s", call_proc_sql)
            queryset = query.fetchall()

            if queryset:
                size = 0
                self.logger.info("mysqlStats, queryset len is : %s" % len(queryset))
                for item in queryset:
                    size += 1
                    if size > limitSize:
                        break
                    temp = {
                        "dept": item[0] if item[0] else "",
                        "count": int(item[1]) if item[1] else 0
                    }
                    result.append(temp)
        return result
        
        
class GetProblemCategoryStats(MyResource):

    @pre_request(request, GetProblemCategoryStatsReq)
    def get(self):
        """质控问题按照类别统计
        """
        response = {"data": []}
        app = self.context.getCaseApplication('hospital')
        params = ["branch", "ward", "department", "attending", "startTime", "endTime", "emrName",
                  "category", "problem", "auditType", "caseType", "problemType", "fixed", "start", "size"]
        req = {c: getattr(request, c) for c in params}
        req['withTotal'] = True

        field, pList = app.getDeptPermission(request.operatorId)
        req['pField'] = field
        req['pList'] = pList
        logging.info(f'user id: {request.operatorId}, permType: {field}, departments: {pList}')
        if request.itemsId:
            req['qcItemsId'] = [0]
            for item in request.itemsId.split(','):
                try:
                    req['qcItemsId'].append(int(item))
                except Exception as e:
                    logging.info(e)
        req['itemType'] = request.itemtype

        req = GetProblemStatsRequest(**req)
        results, total = app.getProblemStats(req)
        response["total"] = total

        qcitems = app.getQcItems()
        for stats in results:
            qcItemId = stats.get('qcItemId')
            counting = stats.get('counting')
            for item in qcitems:
                if item.id == qcItemId:
                    protoItem = {}  # response.data.add()
                    protoItem["id"] = qcItemId
                    protoItem["instruction"] = item.requirement or ''
                    protoItem["emrName"] = item.standard_emr if item.standard_emr and item.standard_emr != '0' else '缺失文书'
                    protoItem["category"] = QCITEM_CATEGORY.get(item.category, "")
                    protoItem["count"] = counting
                    response["data"].append(protoItem)
                    break
        return response


class GetCaseByProblem(MyResource):

    @pre_request(request, GetProblemCategoryStatsReq)
    def get(self):
        """查询存在某个质控问题的病历列表
        """
        response = {"data": []}
        app = self.context.getCaseApplication('hospital')
        params = ["branch", "ward", "department", "attending", "startTime", "endTime",
                  "auditType", "caseType", "problemType", "fixed", "start", "size"]
        req = {c: getattr(request, c) for c in params}
        req['qcItemsId'] = [request.itemId]
        req['withTotal'] = True

        field, pList = app.getDeptPermission(request.operatorId)
        req['pField'] = field
        req['pList'] = pList
        logging.info(f'user id: {request.operatorId}, permType: {field}, departments: {pList}')

        req = GetProblemStatsRequest(**req)

        results, total, case_pstatus, problem_data = app.getProblemStatsCase(req)
        response["total"] = total
        pstatus = {item.get('caseId'): item.get('pstatus') for item in case_pstatus}
        for item in results:
            protoItem = {"caseTags": []}  # response.data.add()
            protoItem["id"] = item.id
            protoItem["caseId"] = item.caseId or ""
            protoItem["patientId"] = item.inpNo or item.patientId or ""
            protoItem["visitTimes"] = item.visitTimes or 0
            protoItem["name"] = item.name or ""
            protoItem["age"] = str(item.age or 0) + str(item.ageUnit or "")
            protoItem["gender"] = item.gender or ""
            protoItem["branch"] = item.branch or ""
            protoItem["department"] = item.department or ""
            protoItem["dischargeDept"] = item.outDeptName or ""
            protoItem["attendDoctor"] = item.attendDoctor or ""
            protoItem["admitTime"] = item.admitTime.strftime('%Y-%m-%d') if item.admitTime else ""
            protoItem["dischargeTime"] = item.dischargeTime.strftime('%Y-%m-%d') if item.dischargeTime else ""
            protoItem["status"] = item.status or 0
            protoItem["statusName"] = parseCaseStatus(protoItem["status"])
            protoItem["ward"] = item.wardName or ""
            if item.TagsModel:
                for tag in item.TagsModel:
                    tagItem = {}  # protoItem.caseTags.add()
                    tagItem["name"] = tag.name or ""
                    tagItem["code"] = tag.code
                    tagItem["icon"] = tag.icon.decode() or ""
                    protoItem["caseTags"].append(tagItem)
            protoItem["fixed"] = pstatus.get(item.caseId)
            problem_info = problem_data.get(item.caseId, {})
            protoItem["fromAi"] = AI_DICT.get(problem_info.get("from_ai") or 0, "未知")
            protoItem["createUser"] = problem_info.get("create_doctor") or ""
            protoItem["createTime"] = problem_info.get("create_time") or ""
            protoItem["solveUser"] = problem_info.get("fix_doctor") or ""
            protoItem["solveTime"] = problem_info.get("fix_time") or ""
            response["data"].append(protoItem)
        return response


class ExportProblemCateStats(MyResource):

    @pre_request(request, GetProblemCategoryStatsReq)
    def post(self):
        """质控问题统计导出
        """
        response = {}
        app = self.context.getCaseApplication('hospital')
        params = ["branch", "ward", "department", "attending", "startTime", "endTime", "emrName",
                  "category", "problem", "auditType", "caseType", "problemType", "fixed"]
        req = {c: getattr(request, c) for c in params}
        req['size'] = 1000

        field, pList = app.getDeptPermission(request.operatorId)
        req['pField'] = field
        req['pList'] = pList
        logging.info(f'user id: {request.operatorId}, permType: {field}, departments: {pList}')

        req = GetProblemStatsRequest(**req)
        results, tmp = app.getProblemStats(req)

        data = []
        qcitems = app.getQcItems()
        # print(f'一共[{len(results)}]个质控问题')
        for stats in results:
            qcItemId = stats.get('qcItemId')
            counting = stats.get('counting')
            for item in qcitems:
                if item.id == qcItemId:
                    req = {c: getattr(request, c) for c in params}
                    req['qcItemsId'] = [qcItemId]
                    req['size'] = 1000
                    req['pField'] = field
                    req['pList'] = pList
                    req = GetProblemStatsRequest(**req)
                    # print(f'查询[{item.requirement}]相关病历列表')
                    case_list, total, case_pstatus, _ = app.getProblemStatsCase(req)
                    pstatus = {item.get('caseId'): item.get('pstatus') for item in case_pstatus}
                    for c in case_list:
                        data.append((qcItemId, item.requirement, counting, c.caseId, c.inpNo or c.patientId, c.name, c.department, c.outDeptName, c.admitTime, c.dischargeTime, c.attendDoctor, c.status, pstatus.get(c.caseId, "")))
                    break
        # 排序
        colsTitle = []
        results = []
        patient_id_name = app.app.config.get(Config.QC_PATIENT_ID_NAME)
        if request.exportType == 1:
            colsTitle = ["问题描述", patient_id_name, "姓名", "入院科室", "出院科室", "入院日期", "出院日期", "医生", "病历状态", "问题状态"]
            for item in data:
                results.append((item[1], item[4], item[5], item[6], item[7], item[8], item[9], item[10], parseCaseStatus(item[11]), item[12]))
        if request.exportType == 2:
            data.sort(key=lambda x: (x[3], x[0]))
            colsTitle = [patient_id_name, "姓名", "入院科室", "出院科室", "入院日期", "出院日期", "医生", "病历状态", "问题描述", "问题状态"]
            for item in data:
                results.append((item[4], item[5], item[6], item[7], item[8], item[9], item[10], parseCaseStatus(item[11]), item[1], item[12]))
        # 保存文件
        wb = Workbook()
        ws = wb.active
        ws.append(tuple(colsTitle))
        bold_font = Font(bold=True)
        for cell in ws["1:1"]:
            cell.font = bold_font
        for row_queryset in results:
            ws.append(row_queryset)
        fileId = QCUtil.getFileId()
        wb.save(os.path.join(export_path, str(fileId) + ".xlsx"))

        response["id"] = fileId
        auditTypeDict = {
            'active': "事中质控",
            'final': "事后质控",
            'department': "科室质控",
            'hospital': "病案质控",
            'firstpage': "编码质控",
            'expert': "专家质控"
        }
        response["fileName"] = datetime.now().strftime('%Y-%m-%d') + "_" + "质控问题统计_" + auditTypeDict.get(request.auditType, '') + ".xlsx"
        return response


class ExpertAllNum(MyResource):

    @pre_request(request, ExpertAllNumReq)
    def post(self):
        """
        专家统计-全院成绩统计-全院质控数量概览
        :param request:
        :param context:
        :return:
        """
        response = {"data": []}
        call_proc_sql = QCUtil.get_call_proc_sql('全院质控数据统计', request, sixMonth=1)
        qc_stats_level = self.context.app.config.get(Config.QC_STATS_LEVEL)
        qc_stats_level = qc_stats_level.split(",")
        level = 2
        if request.level in qc_stats_level:
            level += qc_stats_level.index(request.level) + 2
        QCUtil.get_common_stats_pic_response(self.app, call_proc_sql, response, request, level=level)

        return response


class ExpertAllLevel(MyResource):

    @pre_request(request, ExpertAllNumReq)
    def post(self):
        """
        专家统计-全院成绩统计-质控情况分布
        :param request:
        :param context:
        :return:
        """
        response = {"data": []}
        call_proc_sql = QCUtil.get_call_proc_sql('质控情况分布', request)
        qc_stats_level = self.context.app.config.get(Config.QC_STATS_LEVEL)
        qc_stats_level = qc_stats_level.split(",")
        with self.context.app.mysqlConnection.session() as cursor:
            query = cursor.execute(call_proc_sql)
            self.logger.info("ExpertAllLevel, call_proc_sql: %s", call_proc_sql)
            queryset = query.fetchall()
            total = 0
            first = 0
            second = 0
            third = 0
            for item in queryset:
                total += item[1] or 0
                first += item[3] or 0
                second += item[4] or 0
                third += item[5] or 0
            data = [first, second, third]
            response["total"] = total
            for i in range(len(data)):
                chartData = {}  # response.data.add()
                chartData["xData"] = qc_stats_level[i]
                chartData["yData"] = QCUtil.keep_one(data[i] / total * 100) if total else QCUtil.keep_one(0)
                response["data"].append(chartData)
        return response


class ExpertAllScorePic(MyResource):

    @pre_request(request, ExpertAllNumReq)
    def post(self):
        """
        专家统计-全院成绩统计-全院质控成绩概览折线图
        :param request:
        :param context:
        :return:
        """
        response = {"data": []}
        call_proc_sql = QCUtil.get_call_proc_sql('全院质控数据统计', request, sixMonth=1)
        qc_stats_level = self.context.app.config.get(Config.QC_STATS_LEVEL)
        qc_stats_level = qc_stats_level.split(",")
        level = 3
        if request.level in qc_stats_level:
            level += qc_stats_level.index(request.level) + 1
        QCUtil.get_common_stats_pic_response(self.app, call_proc_sql, response, request, level=level, is_rate=1)

        return response


class ExpertAllDetail(MyResource):

    @pre_request(request, ExpertAllNumReq)
    def get(self):
        """
        专家统计-全院成绩统计-全院质控成绩概览
        :return:
        """
        response = {"headers": [], "data": [], "pageInfo": {}}
        target_name = "质控情况分布"
        to_web_data, data_yaml, total = QCUtil.get_common_header_data(self.app, target_name, request)

        cfg = BIFormConfig.fromYaml(data_yaml)
        processor = BIDataProcess(cfg, to_web_data)
        header, result = processor.toWeb(sortBy=[("总数", -1)], start=0, size=0)
        QCUtil.format_header_data(response, header, result)

        response["pageInfo"]["total"] = total
        return response


class ExpertAllDetailExport(MyResource):

    @pre_request(request, ExpertAllNumReq)
    def get(self):
        """
        专家统计-全院成绩统计-全院质控成绩概览导出
        :param request:
        :param context:
        :return:
        """
        response = {}
        target_name = "质控情况分布"
        to_excel_data, data_yaml, total = QCUtil.get_common_header_data(self.app, target_name, request)

        file_name = "全院质控成绩统计_{}".format(datetime.now().strftime("%Y-%m-%d")) + ".xlsx"
        file_id = self.get_file_id(file_name)
        path_file_name = os.path.join(export_path, file_id + ".xlsx")

        cfg = BIFormConfig.fromYaml(data_yaml)
        processor = BIDataProcess(cfg, to_excel_data)
        processor.toExcel(path=path_file_name, sheet_name=target_name, sortBy=[("总数", -1)])

        response["id"] = file_id
        response["fileName"] = file_name
        return response


class ExpertDeptScorePic(MyResource):

    @pre_request(request, ExpertAllNumReq)
    def post(self):
        """
        专家统计-科室成绩统计-内外科成绩图
        :return:
        """
        response = {"data": []}
        qc_stats_level = self.context.app.config.get(Config.QC_STATS_LEVEL)
        qc_stats_level = qc_stats_level.split(",")
        target_name = "内外科科室成绩"
        call_proc_sql = QCUtil.get_call_proc_sql(target_name, request)
        self.logger.info("ExpertDeptScorePic, call_proc_sql: %s", call_proc_sql)
        data = {}
        with self.context.app.mysqlConnection.session() as cursor:
            query = cursor.execute(call_proc_sql)
            queryset = query.fetchall()
            for item in queryset:
                if item[0]:
                    if not data.get(item[0], []):
                        data[item[0]] = []
                    level = 3
                    if request.level in qc_stats_level:
                        level += qc_stats_level.index(request.level) + 1
                    level_data = QCUtil.keep_one(item[level])
                    if level != 3:  # 甲乙丙率
                        level_data = QCUtil.keep_one(item[level] / item[2] *  100)
                    data[item[0]].append({"department": item[1] or "", "total": item[2] or 0,
                                          "data": level_data})
        for key, value in data.items():
            protoItem = {"items": []}  # response.data.add()
            protoItem["dataName"] = key
            for item in value:
                detailItem = {}  # protoItem.items.add()
                detailItem["xData"] = item["department"]
                detailItem["yData"] = item["data"]
                protoItem["items"].append(detailItem)
            response["data"].append(protoItem)
        return response


class ExpertDeptScoreLevel(MyResource):

    @pre_request(request, ExpertAllNumReq)
    def post(self):
        """
        专家统计-科室成绩统计-内外科成绩级别概览
        :param request:
        :param context:
        :return:
        """
        response = {"items": []}
        qc_stats_level = self.context.app.config.get(Config.QC_STATS_LEVEL)
        qc_stats_level = qc_stats_level.split(",")
        level_list = ["first", "second", "third"]
        qc_stats_level_dict = {qc_stats_level[i]: level_list[i] for i in range(len(qc_stats_level))}
        target_name = "内外科科室成绩"
        call_proc_sql = QCUtil.get_call_proc_sql(target_name, request)
        self.logger.info("ExpertDeptScorePic, call_proc_sql: %s", call_proc_sql)
        data = {"内科": {"total": 0, "first": 0, "second": 0, "third": 0}, "外科": {"total": 0, "first": 0, "second": 0, "third": 0}}
        with self.context.app.mysqlConnection.session() as cursor:
            query = cursor.execute(call_proc_sql)
            queryset = query.fetchall()
            for item in queryset:
                if item[0]:
                    data[item[0]]["total"] += int(item[2] or 0)
                    data[item[0]]["first"] += int(item[4] or 0)
                    data[item[0]]["second"] += int(item[5] or 0)
                    data[item[0]]["third"] += int(item[6] or 0)
        for key, value in data.items():
            protoItem = {"data": {"data": []}}  # response.items.add()
            protoItem["deptType"] = key
            protoItem["data"]["total"] = value["total"]
            for item in qc_stats_level:
                levelItem = {}  # protoItem.data.data.add()
                levelItem["xData"] = item
                levelItem["yData"] = str(value[qc_stats_level_dict[item]])
                protoItem["data"]["data"].append(levelItem)
            response["items"].append(protoItem)

        return response


class ExpertDeptScoreList(MyResource):

    @pre_request(request, ExpertAllNumReq)
    def get(self):
        """
        专家统计-科室成绩统计-内外科成绩列表
        :param request:
        :param context:
        :return:
        """
        response = {"headers": [], "data": [], "pageInfo": {}}
        target_name = "科室成绩概览"
        to_web_data, data_yaml, total = QCUtil.get_common_header_data(self.app, target_name, request)

        cfg = BIFormConfig.fromYaml(data_yaml)
        processor = BIDataProcess(cfg, to_web_data)
        sort_by = [("平均分", -1)]
        if request.sortBy:
            sort_by = [(request.sortBy, request.sortWay)]
        header, result = processor.toWeb(sortBy=sort_by, start=0, size=0)
        QCUtil.format_header_data(response, header, result, not_sort_header=EXPERT_NO_SORT_FIELD)

        response["pageInfo"]["total"] = total
        response["pageInfo"]["start"] = request.start or 0
        response["pageInfo"]["size"] = request.size or 15
        return response

        
class ExpertDeptScoreDetail(MyResource):

    @pre_request(request, ExpertAllNumReq)
    def get(self):
        """
        专家统计-科室成绩统计-内外科成绩列表内科室详情
        :param request:
        :param context:
        :return:
        """
        response = {"picData": {"data": []}, "detailData": {"headers": [], "data": [], "pageInfo": {}}}
        target_name = "科室成绩概览月份数据"
        qc_stats_level = self.context.app.config.get(Config.QC_STATS_LEVEL)
        qc_stats_level = qc_stats_level.split(",")
        level_dict = {"平均分": 2}
        index = 3
        for item in qc_stats_level:
            level_dict[item] = index
            index += 1
        call_proc_sql = QCUtil.get_call_proc_sql(target_name, request)
        self.logger.info("ExpertDeptScoreDetail, call_proc_sql: %s", call_proc_sql)
        data = []
        total = 0
        data_yaml = EXPERT_DEPARTMENT_DETAIL_YAML.replace(
            "[first]", qc_stats_level[0]).replace("[second]", qc_stats_level[1]).replace("[third]", qc_stats_level[2])
        with self.context.app.mysqlConnection.session() as cursor:
            query = cursor.execute(call_proc_sql)
            queryset = query.fetchall()
            for level in level_dict:
                pic_data = {"items": []}  # response.picData.data.add()
                pic_data["dataName"] = level
                index = level_dict[level]
                month_list = []
                for item in queryset:
                    month_list.append(item[0])
                month_list.sort()
                for month in month_list:
                    for item in queryset:
                        if item[0] == month:
                            pic_items = {}  # pic_data.items.add()
                            pic_items.xData = item[0]
                            pic_items.yData = str(item[index])
                            pic_data["items"].append(pic_items)
                response["picData"]["data"].append(pic_data)
            for item in queryset:
                tmp = {"时间": item[0], "总数": int(item[1]), "平均分": QCUtil.keep_one(item[2]), "%s数" % qc_stats_level[0]: int(item[3]),
                       "%s数" % qc_stats_level[1]: int(item[4]),
                       "%s数" % qc_stats_level[2]: int(item[5]), "sort_time": item[0].replace("-", ""),
                       "%s率" % qc_stats_level[0]: QCUtil.keep_one(int(item[3]) / int(item[1]) * 100, True),
                       "%s率" % qc_stats_level[1]: QCUtil.keep_one(int(item[4]) / int(item[1]) * 100, True),
                       "%s率" % qc_stats_level[2]: QCUtil.keep_one(int(item[5]) / int(item[1]) * 100, True)}
                data.append(tmp)
                total += 1
            if not queryset:
                append_dict = dict(EXPERT_DEPARTMENT_DETAIL_DATA, **EXPERT_LEVEL_DATA_1) if StatsRepository.verify_stats_level(qc_stats_level) \
                    else dict(EXPERT_DEPARTMENT_DETAIL_DATA, **EXPERT_LEVEL_DATA_2)
                data.append(append_dict)

        cfg = BIFormConfig.fromYaml(data_yaml)
        processor = BIDataProcess(cfg, data)
        sort_by = []
        if request.sortBy:
            field = request.sortBy if request.sortBy != "时间" else "sort_time"
            sort_by = [(field, request.sortWay)]
        header, result = processor.toWeb(sortBy=sort_by, start=0, size=0)
        QCUtil.format_header_data(response["detailData"], header, result, not_sort_header=EXPERT_NO_SORT_FIELD)
        response["detailData"]["pageInfo"]["total"] = total
        return response


class ExpertDeptScoreDetailExport(MyResource):

    @pre_request(request, ExpertAllNumReq)
    def get(self):
        """
        专家统计-科室成绩统计-内外科成绩列表导出
        :param request:
        :param response:
        :return:
        """
        response = {}
        target_name = "科室成绩概览"
        to_excel_data, data_yaml, total = QCUtil.get_common_header_data(self.app, target_name, request)
        sort_by = []
        if request.sortBy:
            sort_by = [(request.sortBy, request.sortWay)]

        file_name = "科室成绩统计_{}".format(datetime.now().strftime("%Y-%m-%d")) + ".xlsx"
        file_id = QCUtil.get_file_id(file_name)
        path_file_name = os.path.join(export_path, file_id + ".xlsx")

        cfg = BIFormConfig.fromYaml(data_yaml)
        processor = BIDataProcess(cfg, to_excel_data)
        processor.toExcel(path=path_file_name, sortBy=sort_by, sheet_name="科室成绩统计")

        response["id"] = file_id
        response["fileName"] = file_name
        return response


class ExpertDoctorScore(MyResource):

    @pre_request(request, ExpertAllNumReq)
    def get(self):
        """
        专家统计-医生成绩统计-查询
        :param request:
        :param response:
        :return:
        """
        response = {"headers": [], "data": [], "pageInfo": {}}
        start = request.start or 0
        size = request.size or 15
        target_name = "医生成绩统计"
        to_web_data, data_yaml, total = QCUtil.get_common_header_data(self.app, target_name, request)

        cfg = BIFormConfig.fromYaml(data_yaml)
        processor = BIDataProcess(cfg, to_web_data)
        sort_by = [("平均分", -1)]
        if request.sortBy:
            sort_by = [(request.sortBy, request.sortWay)]
        header, result = processor.toWeb(sortBy=sort_by, start=start, size=size)
        QCUtil.format_header_data(response, header, result)

        response["pageInfo"]["total"] = total
        response["pageInfo"]["start"] = start
        response["pageInfo"]["size"] = size
        return response

        
class ExpertDoctorScoreExport(MyResource):

    @pre_request(request, ExpertAllNumReq)
    def get(self):
        """
        专家统计-医生成绩统计-导出
        :param request:
        :param context:
        :return:
        """
        response = {}
        target_name = "医生成绩统计"
        to_excel_data, data_yaml, total = QCUtil.get_common_header_data(self.app, target_name, request)

        sort_by = []
        if request.sortBy:
            sort_by = [(request.sortBy, request.sortWay)]

        file_name = "医生成绩统计_{}".format(datetime.now().strftime("%Y-%m-%d")) + ".xlsx"
        file_id = QCUtil.get_file_id(file_name)
        path_file_name = os.path.join(export_path, file_id + ".xlsx")

        cfg = BIFormConfig.fromYaml(data_yaml)
        processor = BIDataProcess(cfg, to_excel_data)
        processor.toExcel(path=path_file_name, sortBy=sort_by, sheet_name=target_name)

        response["id"] = file_id
        response["fileName"] = file_name
        return response


class GetUpdateTime(MyResource):

    def get(self):
        """
        专家统计-获取更新时间
        :param request:
        :param context:
        :return:
        """
        response = {}
        result = QCUtil.getLastUpdateTime(self.app, "expert")
        updateDatetime = result.get("updateTime", None)
        self.logger.info("lastUpdateTime: %s", updateDatetime)
        if updateDatetime:
            response["lastUpdateTime"] = updateDatetime.strftime('%Y-%m-%d %H:%M:%S') if updateDatetime else ""
        return response
        
        
class StatsDefectRateList(MyResource):

    @pre_request(request, StatsDefectRateListReq)
    def get(self):
        """
        质控分析-缺陷率统计-查询
        :return:
        """
        response = {"data": []}
        if not request.type or request.type not in ("科室", "病区"):
            self.logger.error("StatsDefectRateList, request.type: %s is error.", request.type)
            return get_error_resp("request.type: %s is error." % request.type)
        start = request.start or 0
        size = request.size or 10
        is_need_group = int(self.context.app.config.get(Config.QC_CASE_GROUP_FLAG) or 0)
        data = QCUtil.get_defect_rate_data(self.app, request, is_need_group)
        response["total"] = len(data)
        response["start"] = start
        response["size"] = size
        for item in data[start: start + size]:
            protoItem = {}  # response.data.add()
            protoItem["department"] = item.get("dept") or ""
            protoItem["ward"] = item.get("dept") or ""
            protoItem["group"] = item.get("group") or ""
            protoItem["attend"] = item.get("attend") or ""
            protoItem["applyCaseNum"] = item.get("apply_num") or 0
            protoItem["applyUnqualifiedNum"] = item.get("defect_num") or 0
            protoItem["defectRate"] = QCUtil.keep_one(protoItem["applyUnqualifiedNum"] / protoItem["applyCaseNum"] * 100, True) if protoItem["applyCaseNum"] else "0%"
            protoItem["departmentHide"] = item.get("deptHide") or ""
            protoItem["wardHide"] = item.get("deptHide") or ""
            protoItem["groupHide"] = item.get("groupHide") or ""
            protoItem["attendHide"] = item.get("attendHide") or ""
            response["data"].append(protoItem)

        return response


class StatsDefectRateExport(MyResource):

    @pre_request(request, StatsDefectRateListReq)
    def get(self):
        """
        质控分析-缺陷率统计-导出
        :return:
        """
        response = {}
        is_need_group = int(self.context.app.config.get(Config.QC_CASE_GROUP_FLAG) or 0)
        data = QCUtil.get_defect_rate_data(self.app, request, is_need_group)
        file_name = "{}至{}缺陷率统计".format(request.startTime[:10], request.endTime[:10]) + ".xlsx"
        file_id = QCUtil.get_file_id(file_name)
        path_file_name = os.path.join(export_path, file_id + ".xlsx")
        if request.type == "科室":
            if is_need_group == 1:
                title = DEFECT_RATE_TITLE_DEPT
            else:
                title = DEFECT_RATE_TITLE_DEPT_NO_GROUP
        else:
            title = DEFECT_RATE_TITLE_WARD
        StatsRepository.write_defect_rate_excel(title, data, path_file_name)
        response["id"] = file_id
        response["fileName"] = file_name
        return response


class StatsDefectRateDetailList(MyResource):

    @pre_request(request, StatsDefectRateListReq)
    def get(self):
        """
        质控分析-缺陷率统计-明细-查询
        :return:
        """
        response = {"data": []}
        case_model = self.context.app.mysqlConnection["case"]
        with self.context.app.mysqlConnection.session() as session:
            total, data = StatsRepository.get_defect_rate_detail_data(session, case_model, request)
            response["total"] = total
            response["start"] = request.start or 0
            response["size"] = request.size or 10

            for item, case in data:
                protoItem = {}  # response.data.add()
                protoItem["caseId"] = item.caseid or ""
                protoItem["patientId"] = case.inpNo or case.patientId or ""
                protoItem["name"] = item.name or ""
                protoItem["admitTime"] = item.admittime.strftime("%Y-%m-%d") if item.admittime else ""
                protoItem["dischargeTime"] = item.dischargetime.strftime("%Y-%m-%d") if item.dischargetime else ""
                protoItem["department"] = item.outdeptname or ""
                protoItem["group"] = item.medicalgroupname or ""
                protoItem["ward"] = item.outhosward or ""
                protoItem["attend"] = item.attendDoctor or ""
                protoItem["score"] = str(item.score or 0)
                protoItem["isQualified"] = item.is_standard or ""
                protoItem["status"] = item.status or 0
                protoItem["statusName"] = parseStatusName(False, item.status)

        return response


class StatsDefectRateDetailExport(MyResource):

    @pre_request(request, StatsDefectRateListReq)
    def get(self):
        """
        质控分析-缺陷率统计-明细-导出
        :return:
        """
        response = {}
        is_need_group = int(self.context.app.config.get(Config.QC_CASE_GROUP_FLAG) or 0)
        case_model = self.context.app.mysqlConnection["case"]
        with self.context.app.mysqlConnection.session() as session:
            total, data = StatsRepository.get_defect_rate_detail_data(session, case_model, request, is_export=1)
            if request.type == "科室":
                if is_need_group == 1:
                    title = DEFECT_RATE_DETAIL_TITLE_DEPT
                else:
                    title = DEFECT_RATE_DETAIL_TITLE_DEPT_NO_GROUP
            else:
                title = DEFECT_RATE_DETAIL_TITLE_WARD
            title_name = StatsRepository.get_defect_rate_detail_export_title_name(request)
            file_name = "{}至{}【{}】明细".format(request.startTime[:10], request.endTime[:10], title_name) + ".xlsx"
            file_id = QCUtil.get_file_id(file_name)
            path_file_name = os.path.join(export_path, file_id + ".xlsx")
            StatsRepository.write_defect_detail_excel(title, data, path_file_name)
            response["id"] = file_id
            response["fileName"] = file_name
        return response


class StatsDefectRateUpdateStatus(MyResource):

    def get(self):
        """
        质控分析-缺陷率统计-更新状态查询
        :param request:
        :param context:
        :return:
        """
        response = {}
        with self.context.app.mysqlConnection.session() as session:
            query_sql = '''select updatetime, updatestatus from case_updatestatus_rate'''
            query = session.execute(query_sql)
            data = query.fetchone()
            response["status"] = int(data[1] if data else 2)
            last_time = data[0] + timedelta(hours=8) if data else datetime(year=2022, month=1, day=1)
            response["lastUpdateTime"] = last_time.strftime('%Y-%m-%d %H:%M:%S') if data else ""
        return response


class StatsDefectRateUpdate(MyResource):

    def get(self):
        """
        质控分析-缺陷率统计-更新数据
        :param request:
        :param context:
        :return:
        """
        with self.context.app.mysqlConnection.session() as session:
            call_sql = '''call pr_case_rate_process('','') '''
            session.execute(call_sql)
            session.commit()
        return g.result


class StatsArchivedQualityList(MyResource):

    @pre_request(request, StatsDefectRateListReq)
    def get(self):
        """质控分析-归档病历质量统计-查询
        """
        response = {"data": []}
        if not request.type or request.type not in ("科室", "病区"):
            self.logger.error("StatsDefectRateList, request.type: %s is error.", request.type)
            return get_error_resp("request.type: %s is error." % request.type)
        start = request.start or 0
        size = request.size or 10

        data = []
        with self.context.app.mysqlConnection.session() as session:
            group_flag = int(self.context.app.config.get(Config.QC_CASE_GROUP_FLAG) or 0) == 1
            data = ArchivedQualityStats(request.type, request.branch, request.department, request.group, 
                                        request.ward, request.attend, request.startTime, 
                                        request.endTime).stats(session, group_flag)

        response["total"] = len(data)
        response["start"] = start
        response["size"] = size
        for item in data[start: start + size]:
            protoItem = {}  # response.data.add()
            if item.attend:
                protoItem["attend"] = item.attend or ""
            elif item.group:
                protoItem["group"] = item.group or ""
            else:
                protoItem["department"] = item.dept or ""
                protoItem["ward"] = item.ward or ""
            protoItem["archivedNum"] = int(item.data[0]) or 0
            protoItem["finishedNum"] = int(item.data[1]) or 0
            protoItem["averageScore"] = self.keep_one(item.data[2])
            protoItem["sampleRate"] = self.keep_one(item.data[3] * 100, True)
            protoItem["firstNum"] = int(item.data[4])
            protoItem["firstAvgScore"] = self.keep_one(item.data[5])
            protoItem["firstRate"] = self.keep_one(item.data[6] * 100, True)
            protoItem["secondNum"] = int(item.data[7])
            protoItem["secondAvgScore"] = self.keep_one(item.data[8])
            protoItem["secondRate"] = self.keep_one(item.data[9] * 100, True)
            protoItem["thirdNum"] = int(item.data[10])
            protoItem["thirdAvgScore"] = self.keep_one(item.data[11])
            protoItem["thirdRate"] = self.keep_one(item.data[12] * 100, True)
            protoItem["departmentHide"] = req.department if item.dept == "总计" else (item.dept or "")
            protoItem["wardHide"] = req.ward if item.ward == "总计" else (item.ward or "")
            protoItem["groupHide"] = item.group or req.group or ""
            protoItem["attendHide"] = item.attend or req.attend or ""
            response["data"].append(protoItem)
        return response


class StatsArchivedQualityExport(MyResource):

    @pre_request(request, StatsDefectRateListReq)
    def get(self):
        """
        质控分析-归档病历质量统计-导出
        """
        response = {}
        file_name = "{}至{}归档病历质量统计".format(request.startTime[:10], request.endTime[:10]) + ".xlsx"
        file_id = QCUtil.get_file_id(file_name)
        path_file_name = os.path.join(export_path, file_id + ".xlsx")

        header = ["科室", "诊疗组", "责任医生", "已归档病历数", "质控病历数", "总分", "甲级病历数", "甲级病历总分", "乙级病历数", "乙级病历总分", "丙级病历数", "丙级病历总分"]
        if request.type == "科室" and int(self.context.app.config.get(Config.QC_CASE_GROUP_FLAG) or 0) == 1:
            header = ["科室", "责任医生", "已归档病历数", "质控病历数", "总分", "甲级病历数", "甲级病历总分", "乙级病历数", "乙级病历总分", "丙级病历数", "丙级病历总分"]
        if request.type == "病区":
            header = ["病区", "责任医生", "已归档病历数", "质控病历数", "总分", "甲级病历数", "甲级病历总分", "乙级病历数", "乙级病历总分", "丙级病历数", "丙级病历总分"]

        with self.context.app.mysqlConnection.session() as session:
            stats = ArchivedQualityStats(request.type, request.branch, request.department, request.group, request.ward, request.attend, request.startTime, request.endTime)
            stats.export_excel(session, header, path_file_name)

        response["id"] = file_id
        response["fileName"] = file_name
        return response

        
class StatsArchivedQualityDetailList(MyResource):

    @pre_request(request, StatsDefectRateListReq)
    def get(self):
        """
        质控分析-归档病历质量统计-明细-查询
        :return:
        """
        response = {"data": []}
        case_model = self.context.app.mysqlConnection["case"]
        with self.context.app.mysqlConnection.session() as session:
            stats = ArchivedQualityStats(request.type, request.branch, request.department, request.group, 
                                         request.ward, request.attend, request.startTime, request.endTime)
            total, data = stats.detail(session, case_model, request)
            response["total"] = total
            response["start"] = request.start or 0
            response["size"] = request.size or 10

            for item, case in data:
                protoItem = {}  # response.data.add()
                protoItem["caseId"] = item.caseid or ""
                protoItem["patientId"] = case.inpNo or case.patientId or ""
                protoItem["name"] = item.name or ""
                protoItem["admitTime"] = item.admittime.strftime("%Y-%m-%d") if item.admittime else ""
                protoItem["dischargeTime"] = item.dischargetime.strftime("%Y-%m-%d") if item.dischargetime else ""
                protoItem["department"] = item.outdeptname or ""
                protoItem["group"] = item.medicalgroupname or ""
                protoItem["ward"] = item.wardname or ""
                protoItem["attend"] = item.attendDoctor or ""
                protoItem["score"] = str(item.score or 0)
                protoItem["isFinished"] = item.is_mq or ""
                protoItem["level"] = {"甲": "甲级", "乙": "乙级", "丙": "丙级"}.get(item.caselevel, item.caselevel) or ""
                response["data"].append(protoItem)

        return response


class StatsArchivedQualityDetailExport(MyResource):

    @pre_request(request, StatsDefectRateListReq)
    def get(self):
        """
        质控分析-归档病历质量统计-明细-导出
        :return:
        """
        response = {}
        case_model = self.context.app.mysqlConnection["case"]
        with self.context.app.mysqlConnection.session() as session:
            title_name = StatsRepository.get_defect_rate_detail_export_title_name(request)
            file_name = "{}至{}【{}】明细".format(request.startTime[:10], request.endTime[:10], title_name) + ".xlsx"
            file_id = QCUtil.get_file_id(file_name)
            path_file_name = os.path.join(export_path, file_id + ".xlsx")

            stats = ArchivedQualityStats(request.type, request.branch, request.department, request.group, 
                                         request.ward, request.attend, request.startTime, request.endTime)
            total, data = stats.detail(session, case_model, request, is_export=True)

            header = ["病历号", "姓名", "入院日期", "出院日期", "科室", "责任医生", "是否质控", "病历等级", "病历分数"]
            if request.type == "科室" and int(self.context.app.config.get(Config.QC_CASE_GROUP_FLAG) or 0) == 1:
                header = ["病历号", "姓名", "入院日期", "出院日期", "科室", "诊疗组", "责任医生", "是否质控", "病历等级", "病历分数"]
            if request.type == "病区":
                header = ["病历号", "姓名", "入院日期", "出院日期", "病区", "责任医生", "是否质控", "病历等级", "病历分数"]

            stats.write_detail_excel(header, data, path_file_name)

            response["id"] = file_id
            response["fileName"] = file_name
        return response


class StatsArchivedQualityUpdateStatus(MyResource):

    def get(self):
        """
        质控分析-归档病历质量统计-更新状态查询
        """
        response = {}
        with self.context.app.mysqlConnection.session() as session:
            query_sql = '''select updatetime, updatestatus from case_archive_extend_updatestatus '''
            query = session.execute(query_sql)
            data = query.fetchone()
            response["status"] = int(data[1] if data else 2)
            response["lastUpdateTime"] = arrow.get(data[0] if data else "2022-01-01").to('+08:00').strftime('%Y-%m-%d %H:%M:%S')
        return response


class StatsArchivedQualityUpdate(MyResource):

    def get(self):
        """
        质控分析-归档病历质量统计-更新数据
        """
        with self.context.app.mysqlConnection.session() as session:
            call_sql = '''call pr_case_archivecase_process('','') '''
            session.execute(call_sql)
        return g.result


class StatsRunningCaseNum(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-事中质控情况分析-病历数、累计减少问题数、平均病历减少问题数
        :param request:
        :param context:
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_running_case_info(session, request, response)
            StatsRepository.get_running_problem_info(session, request, response)
        return response


class StatsRunningDeptTop(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-事中质控情况分析-问题改善科室TOP10
        :param request:
        :param context:
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_running_dept_top(session, request, response)
        return response


class StatsRunningDeptInfo(MyResource):
    
    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-事中质控情况分析-对应科室问题改善诊疗组、医生TOP10情况
        :param request:
        :param context:
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        is_need_group = int(self.context.app.config.get(Config.QC_CASE_GROUP_FLAG) or 0)
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_running_dept_info(session, request, response, is_need_group)
        return response


class StatsRunningType(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-事中质控情况分析-各类别问题改善情况分析
        :param request:
        :param context:
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_running_type(session, request, response)
        return response


class StatsRunningTypeInfo(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-事中质控情况分析-对应类别问题改善情况分析
        :param request:
        :param context:
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_running_type_info(session, request, response)
        return response

        
class StatsVetoBaseInfo(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-强制拦截情况分析-病历拦截率、强制拦截数、累计减少问题数
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_veto_base_info(session, request, response)
        return response


class StatsVetoCaseTrendInfo(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-强制拦截情况分析-病历强制拦截率趋势分析、累计减少强制问题数趋势分析
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_veto_case_trend_info(session, request, response)
        return response


class StatsVetoDeptTopInfo(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-强制拦截情况分析-病历强制拦截率科室top10
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_veto_dept_top_info(session, request, response)
        return response


class StatsVetoDoctorTopInfo(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-强制拦截情况分析-对应科室医生top10
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        is_need_group = int(self.context.app.config.get(Config.QC_CASE_GROUP_FLAG) or 0)
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_veto_doctor_top_info(session, request, response, is_need_group)
        return response


class StatsVetoProblemTypeInfo(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-强制拦截情况分析-问题所属类别分析
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_veto_problem_type_info(session, request, response)
        return response


class StatsVetoProblemNumInfo(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-强制拦截情况分析-对应问题所属类别数量分析
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_veto_problem_num_info(session, request, response)
        return response


class StatsRefuseCaseNumInfo(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-病历退回情况分析-退回病历数、退回率
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_refuse_case_num_info(session, request, response)
        return response


class StatsRefuseRatioInfo(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-病历退回情况分析-病历退回率、退回病历数分析
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_refuse_ratio_info(session, request, response)
        return response


class StatsRefuseDeptTopInfo(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-病历退回情况分析-病历退回科室top10
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_refuse_dept_top_info(session, request, response)
        return response


class StatsRefuseDoctorTopInfo(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-病历退回情况分析-对应科室重点关注医生top10
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        is_need_group = int(self.context.app.config.get(Config.QC_CASE_GROUP_FLAG) or 0)
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_refuse_doctor_top_info(session, request, response, is_need_group)
        return response


class StatsRefuseProblemTypeInfo(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-病历退回情况分析-问题所属类别分析
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_refuse_problem_type_info(session, request, response)
        return response


class StatsRefuseProblemNumInfo(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-病历退回情况分析-问题触发数量分析
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_refuse_problem_num_info(session, request, response)
        return response


class StatsArchiveCaseNumInfo(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-事后质控情况分析-病历数、病历缺陷率、平均每病案缺陷数
        :param request:
        :param context:
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_archive_case_num(session, request, response)
        return response


class StatsArchiveRatioInfo(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-事后质控情况分析-病历质量趋势分析、病历问题数量趋势分析
        :param request:
        :param context:
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_archive_ratio_info(session, request, response)
        return response


class StatsArchiveDeptTopInfo(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-事后质控情况分析-病历质量重点关注科室top10
        :param request:
        :param context:
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_archive_dept_top_info(session, request, response)
        return response
        

class StatsArchiveDoctorTopInfo(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-事后质控情况分析-对应科室-病历质量重点关注医生top10
        :param request:
        :param context:
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        is_need_group = int(self.context.app.config.get(Config.QC_CASE_GROUP_FLAG) or 0) == 1
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_archive_doctor_top_info(session, request, response, is_need_group)
        return response


class StatsArchiveProblemNumTopInfo(MyResource):
    
    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-事后质控情况分析-病历问题数量重点关注科室top10
        :param request:
        :param context:
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_archive_problem_num_top(session, request, response)
        return response


class StatsArchiveProblemNumDoctorTopInfo(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-事后质控情况分析-对应科室-问题数量重点关注医生top10
        :param request:
        :param context:
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        is_need_group = int(self.context.app.config.get(Config.QC_CASE_GROUP_FLAG) or 0) == 1
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_archive_problem_num_doctor_top(session, request, response, is_need_group)
        return response


class StatsArchiveProblemTypeInfo(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-事后质控情况分析-问题所属类别分析
        :param request:
        :param context:
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_archive_problem_type(session, request, response)
        return response


class StatsArchiveProblemNumInfo(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def get(self):
        """
        质控分析-事后质控情况分析-问题触发数量分析
        :param request:
        :param context:
        :return:
        """
        response = {"targetInfo": [], "targetData": []}
        with self.context.app.mysqlConnection.session() as session:
            StatsRepository.get_archive_problem_num_info(session, request, response)
        return response


class GetWorkloadReport(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def post(self):
        """工作量统计报表
        """
        response = {"headers": [], "data": [], "pageInfo": {}}
        args = [request.branch or "全部院区", request.department or "全部科室", request.startTime, request.endTime]
        title = ["科室", "出院人数", "提交质控病案数", "科室质控病案数", "科室质控病案比例", "院级质控病案数", "院级质控病案比例", "首页质控病案数", "首页质控病案比例", "总质控数量", "总体病案质控比例",
                 "甲级病案数", "乙级病案数", "丙级病案数", "甲级率", "输血患者人数", "科室输血病案质控人数", "科室输血病案质控比例", "院级输血病案质控数", "院级输血病案质控比例",
                 "死亡患者人数", "科室死亡病案质控数", "科室死亡病案质控比例", "院级死亡病案质控数", "院级死亡病案质控比例", "超30天患者人数", "科室超30天病案质控数",
                 "科室超30天病案质控比例", "院级超30天病案质控数", "院级超30天病案质控比例"]
        to_web_data = []
        total = 0
        with self.context.app.mysqlConnection.session() as session:
            call_proc_sql = """call pr_case_report('%s', '%s', '%s', '%s')""" % tuple(args)
            query = session.execute(call_proc_sql)
            queryset = query.fetchall()
            for item in queryset:
                row_data = {}
                for index in range(len(title)):
                    if "比例" in title[index]:
                        row_data[title[index]] = item[index] or 0
                    else:
                        row_data[title[index]] = item[index]
                to_web_data.append(row_data)
                total += 1
        cfg = BIFormConfig.fromYaml(WORKLOAD_REPORT_YAML)
        processor = BIDataProcess(cfg, to_web_data)
        header, result = processor.toWeb(start=request.start, size=request.size)
        QCUtil.format_header_data(response, header, result)

        response["pageInfo"]["start"] = request.start or 0
        response["pageInfo"]["size"] = request.size or 10
        response["pageInfo"]["total"] = total
        return response


class ExportWorkloadReport(MyResource):

    @pre_request(request, GetHospitalArchivingRateReq)
    def post(self):
        """工作量报表导出"""
        response = {}
        file_name = "工作量报表_{}".format(datetime.now().strftime("%Y-%m-%d")) + ".xlsx"
        file_id = QCUtil.get_file_id(file_name)
        path_file_name = os.path.join(export_path, file_id + ".xlsx")

        args = [request.branch or "全部院区", request.department or "全部科室", request.startTime, request.endTime]
        to_excel_data = []

        with self.context.app.mysqlConnection.session() as session:
            call_proc_sql = """call pr_case_report('%s', '%s', '%s', '%s')""" % tuple(args)
            query = session.execute(call_proc_sql)
            queryset = query.fetchall()
            title = ["科室", "出院人数", "提交质控病案数", "科室质控病案数", "科室质控病案比例", "院级质控病案数", "院级质控病案比例", "首页质控病案数", "首页质控病案比例", "总质控数量", "总体病案质控比例",
                     "甲级病案数", "乙级病案数", "丙级病案数", "甲级率", "输血患者人数", "科室输血病案质控人数", "科室输血病案质控比例", "院级输血病案质控数", "院级输血病案质控比例",
                     "死亡患者人数", "科室死亡病案质控数", "科室死亡病案质控比例", "院级死亡病案质控数", "院级死亡病案质控比例", "超30天患者人数", "科室超30天病案质控数",
                     "科室超30天病案质控比例", "院级超30天病案质控数", "院级超30天病案质控比例"]
            for item in queryset:
                row_data = {}
                for index in range(len(title)):
                    if "比例" in title[index]:
                        row_data[title[index]] = item[index] or 0
                    else:
                        row_data[title[index]] = item[index]
                to_excel_data.append(row_data)
        cfg = BIFormConfig.fromYaml(WORKLOAD_REPORT_YAML)
        processor = BIDataProcess(cfg, to_excel_data)
        processor.toExcel(path=path_file_name, sheet_name="工作量报表")

        response["id"] = file_id
        response["fileName"] = file_name
        return response

