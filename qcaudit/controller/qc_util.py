from qcaudit.service.protomarshaler import *
from qcaudit.config import Config
from qcaudit.domain.case.case import CaseType
from qcaudit.common.const import *
from qcaudit.domain.req import SortField
from qcaudit.domain.case.req import GetCaseListRequest
from qcaudit.service.sample_archive_job import SampleArchiveJob
from qcaudit.domain.stats.statsrepository import StatsRepository
import threading, uuid, traceback
from datetime import timedelta
from collections import defaultdict
from openpyxl.styles import Alignment, Font
from qcaudit.utils.towebconfig import *
from qcaudit.utils.bidataprocess import *
try:
    from openpyxl.cell import get_column_letter, column_index_from_string
except ImportError:
    from openpyxl.utils import get_column_letter, column_index_from_string


class QCUtil:

    @classmethod
    def get_case_list_req(cls, app, request, is_export=0, field='', p_list=list()):
        """
        查询病历列表条件获取
        :return:
        """
        params = ["branch", "ward", "department", "attend", "rating",
                  "caseId", "patientId", "reviewer", "problemFlag", "patientName",
                  "autoReviewFlag", "firstPageFlag", "start", "size",
                  "auditType", "auditStep", "startTime", "endTime", "caseType", "deptType", "timeType",
                  "diagnosis", "operation", "archiveRating", "refuseCount", "group", "category",
                  "minScore", "maxScore", "minCost", "maxCost", "activeQcNum", "activeManProblemNum", "activeProblemStatus",
                  "minDays", "maxDays"]
        req = {c: getattr(request, c) for c in params if hasattr(request, c)}
        req['pField'] = field
        req['pList'] = p_list
        req['onlineStartTime'] = app.app.config.get(Config.QC_FIRST_ONLINE_PUBLISH_TIMESTAMP, '')
        setRequestStatus(req, request.auditType, request.auditStep, status=request.status)
        req["is_export"] = is_export
        if request.caseType == 'running':
            req['includeCaseTypes'] = [CaseType.ACTIVE]
        elif request.caseType == 'archived':
            req['includeCaseTypes'] = [CaseType.ARCHIVE]
        elif request.caseType == 'Final':
            req['includeCaseTypes'] = [CaseType.FINAL]
        # 排序
        if request.sortField:
            dept = 'outDeptName' if request.auditType != AUDIT_TYPE_ACTIVE else "department"
            # 抽取顺序
            FIELD_MAP = {
                'department': dept,
                'ward': 'wardName',
                'attending': 'attendDoctor',
                'branch': 'branch',
                'problems': CASE_LIST_PROBLEM_COUNT,
                'tags': 'tags',
                'receiveTime': 'receiveTime',
                'admitTime': 'admitTime',
                'auditTime': CASE_LIST_AUDIT_TIME,
                'dischargeTime': 'dischargeTime',
                'problemCount': CASE_LIST_PROBLEM_COUNT,
                'auditDoctor': CASE_LIST_AUDIT_DOCTOR,
                'caseScore': CASE_LIST_CASE_SCORE,
                'activeManProblemNum': 'now_problem_num',
                'activeProblemStatus': 'problem_all_num - no_fix_problem_num',
            }
            req['sortFields'] = []
            if request.auditType == 'department':
                req['sortFields'] = [SortField(field='urgeFlag', way='DESC', table='audit_record')]
            for sf in request.sortField:
                if FIELD_MAP.get(sf["field"]):
                    if sf["field"] == 'receiveTime':
                        sort_field = SortField(field=FIELD_MAP.get(sf["field"], sf["field"]), way=c,
                                               table='audit_record', extParams=sf.get("extParams", []))
                    elif sf["field"] == "problems" or sf["field"] == "problemCount":
                        sort_field = SortField(field=FIELD_MAP.get(sf["field"])[request.auditType], way=sf.get("way", "asc"),
                                               extParams=sf.get("extParams", []))
                    elif sf["field"] == "auditTime":
                        table = 'audit_record' if request.auditType != AUDIT_TYPE_ACTIVE else "active_record"
                        sort_field = SortField(field=FIELD_MAP.get(sf["field"])[request.auditType], way=sf.get("way", "asc"),
                                               table=table, extParams=sf.get("extParams", []))
                    elif sf["field"] == "auditDoctor":
                        table = 'audit_record' if request.auditType != AUDIT_TYPE_ACTIVE else "active_record"
                        sort_field = SortField(field=FIELD_MAP.get(sf["field"])[request.auditType],
                                               way=sf.get("way", "asc"), table=table, extParams=sf.get("extParams", []))
                    elif sf["field"] == "caseScore":
                        sort_field = SortField(field=FIELD_MAP.get(sf["field"])[request.auditType], way=sf.get("way", "asc"),
                                               table='audit_record', extParams=sf.get("extParams", []))
                    else:
                        sort_field = SortField(field=FIELD_MAP.get(sf["field"], sf["field"]), way=sf.get("way", "asc"),
                                               extParams=sf.get("extParams", []))
                    req['sortFields'].append(sort_field)
        else:
            # 默认排序
            sort_field = 'dischargeTime' if request.auditType != AUDIT_TYPE_ACTIVE else 'admitTime'
            # 科室质控将催办的病历置顶
            if request.auditType == 'department':
                req['sortFields'] = [SortField(field='urgeFlag', way='DESC', table='audit_record'), SortField(field=sort_field, way='DESC')]
            else:
                req['sortFields'] = [SortField(field=sort_field, way='DESC')]
        req['isFinal'] = request.auditStep == "recheck"
        if request.assignDoctor:
            req['sampleExpert'] = request.assignDoctor
        if app.app.config.get(Config.QC_PRECONDITION.format(auditType=request.auditType)):
            req["precondition"] = app.app.config.get(Config.QC_PRECONDITION.format(auditType=request.auditType))
        req["not_apply"] = app.app.config.get(Config.QC_NOT_APPLY_AUDIT.format(auditType=request.auditType))
        if app.app.config.get(Config.QC_SAMPLE_STATUS.format(auditType=request.auditType)) == '1':
            req["openSampleFlag"] = True
            req["sampleArchiveFlag"] = app.app.config.get(Config.QC_SAMPLE_ARCHIVE.format(auditType=request.auditType)) == '1'
        if request.tag:
            req['tags'] = [tag for tag in request.tag.split(',') if tag]
        req['timeType'] = int(request.timeType) if request.timeType else 0
        req['visitType'] = request.visitType or 2
        req['fixOvertimeFlag'] = request.fixOvertimeFlag or 0
        req["sampleByTags"] = request.sampleByTags if hasattr(request, "sampleByTags") else request.tags
        if request.overtime:
            req['overtime'] = request.overtime
        req = GetCaseListRequest(**req)
        return req

    @classmethod
    def check_archive_task(cls, context):
        archive_job = SampleArchiveJob(context)
        archive_task = threading.Thread(target=archive_job.task)
        archive_task.start()
        archive_task.join(0)

    @classmethod
    def get_sample_list_req(cls, app, request):
        params = ["branch", "ward", "department", "attend", "status", "rating",
                  "caseId", "patientId", "reviewer", "problemFlag", "patientName",
                  "autoReviewFlag", "firstPageFlag", "start", "size", "auditType",
                  "startTime", "endTime", "caseType", "group", "minDays", "maxDays", "minScore", "maxScore", "minCost", "maxCost"]
        req = {c: getattr(request, c) for c in params}
        # 重点病历标签过滤
        req["sampleByTags"] = request.tags
        req["tags"] = [tag for tag in request.tag.split(',') if tag]

        # if request.status:
        req["status"] = []
        if request.caseType:
            if request.caseType == 'running':
                req['includeCaseTypes'] = [CaseType.ACTIVE]
            elif request.caseType == 'archived':
                req['includeCaseTypes'] = [CaseType.ARCHIVE]
            elif request.caseType == 'Final':
                req['includeCaseTypes'] = [CaseType.FINAL]
        if request.sortField:
            # 抽取顺序
            FIELD_MAP = {
                'department': 'outDeptName',
                'ward': 'wardName',
                'attending': 'attendDoctor',
                'branch': 'branch',
                'problems': CASE_LIST_PROBLEM_COUNT,
                'tags': 'tags',
                'receiveTime': 'receiveTime',
                'admitTime': 'admitTime',
                'dischargeTime': 'dischargeTime',
                "group": "medicalGroupName",
            }
            req['sortFields'] = []
            for sf in request.sortField:
                if FIELD_MAP.get(sf.field):
                    if sf.field == 'receiveTime':
                        sort_field = SortField(field=FIELD_MAP.get(sf.field, sf.field), way=sf.way,
                                               table='audit_record', extParams=sf.extParams)
                    elif sf.field == "problems":
                        sort_field = SortField(field=FIELD_MAP.get(sf.field)[request.auditType], way=sf.way,
                                               extParams=sf.extParams)
                    else:
                        sort_field = SortField(field=FIELD_MAP.get(sf.field, sf.field), way=sf.way,
                                               extParams=sf.extParams)
                    req['sortFields'].append(sort_field)

        # 抽取归档标记，如果允许抽取归档，过滤掉终末病历列表中已归档的病历
        req["sampleArchiveFlag"] = app.app.config.get(Config.QC_SAMPLE_ARCHIVE.format(auditType=request.auditType)) == '1'
        # 抽取列表仅需要住院病历
        req["visitType"] = "2"

        req = GetCaseListRequest(**req)

    @classmethod
    def keepOne(cls, a):
        """
        保留一位小数
        :return:
        """
        return float("%.1f" % a)

    @classmethod
    def get_start_end_time(cls, request):
        """
        将接收时间序列化为起始/终止时间
        :return:
        """
        query_time = request.time or ""
        timeType = request.timeType or ""
        if not query_time or not timeType:
            return None, None
        if timeType == "quarter":
            query_time_list = query_time.split(",")
            if len(query_time_list) == 2:
                month_last_day = cls.get_month_last_day(query_time)
                return "%s-01" % query_time_list[0], "%s-%s" % (query_time_list[1], month_last_day)
            return None, None
        month_last_day = cls.get_month_last_day(query_time)
        startTime = "%s-01-01" % query_time if timeType == "year" else "%s-01" % query_time
        endTime = "%s-12-31" % query_time if timeType == "year" else "%s-%s" % (query_time, month_last_day)
        if startTime > datetime.now().strftime("%Y-%m-%d"):
            return None, None
        return startTime, endTime

    @classmethod
    def get_month_last_day(cls, day):
        """
        获取月份最后一天
        :param day:
        :return:
        """
        year = int(day[:4])
        if int(day[-2:]) in [4, 6, 9, 11]:
            return 30
        elif int(day[-2:]) == 2:
            return 29 if (not year % 4 and year % 100) or not year % 400 else 28
        return 31

    @classmethod
    def getStatsUpdateState(cls, app, is_expert=0):
        """
        查询统计数据更新状态
        :param is_expert:
        :return:
        """
        result = {}
        with app.mysqlConnection.session() as cursor:
            table = "case_updatestatus" if not is_expert else "case_updatestatus_expert"
            query_sql = "select updatetime, updatestatus from %s" % table
            query = cursor.execute(query_sql)
            ret = query.fetchone()
            if ret:
                result["updatetime"] = ret[0] if ret[0] else ""
                result["status"] = int(ret[1]) if ret[1] else 0
        return result

    @classmethod
    def getLastUpdateTime(cls, app, u_type):
        """
        查询上次更新时间
        :return:
        """
        result = {}
        with app.mysqlConnection.session() as cursor:
            table = cls.get_update_status_table(u_type)
            query_last_update_time_sql = "select updatetime, updatestatus from %s" % table
            query = cursor.execute(query_last_update_time_sql)
            # self.logger.info("getLastUpdateTime, query_last_update_time_sql: %s", query_last_update_time_sql)
            ret = query.fetchone()
            if ret:
                result["status"] = int(ret[1] or 0)
                updateDatetime = ret[0] + timedelta(hours=8) if ret[0] else ""
                # self.logger.info("getLastUpdateTime, updateDatetime: %s", updateDatetime)
                result["updateTime"] = updateDatetime
        return result

    @classmethod
    def get_update_status_table(cls, u_type):
        """
        获取更新状态、时间表名
        :return:
        """
        table = "case_updatestatus_qc"
        if u_type in ("hospital", "table"):
            table = "case_updatestatus_qc"
        elif u_type == "archive":
            table = "case_updatestatus"
        elif u_type == "expert":
            table = "case_updatestatus_expert"
        elif u_type in ("running", "veto", "refuse"):
            table = "case_qc_analyse_updatestatus"
        elif u_type == "workload":
            table = "case_report_updatestatus"
        return table

    @classmethod
    def sortDepartmentArchivingRateResult(cls, queryset, sortWay="asc"):
        """
        科室归档率结果排序
        :param queryset:
        :param sortWay:
        :return:
        """
        resultList = [queryset[0]]
        deptDict = defaultdict(list)
        for item in queryset[1:]:
            # 同一科室的排序号相同，根据排序号分组
            deptDict[item[-1]].append(item)

        deptList = list(deptDict.items())

        if sortWay == "desc":
            sortedDept = sorted(deptList, key=lambda x: (x[0] or 0), reverse=True)
        else:
            sortedDept = sorted(deptList, key=lambda x: (x[0] or 0))
        for item in sortedDept:
            if sortWay == "desc":
                sortedRows = sorted(item[1] or "-", key=lambda x: (2 if x[0] and x[0].endswith(u"汇总") else 1), reverse=True)
            else:
                sortedRows = sorted(item[1] or "-", key=lambda x: (0 if x[0] and x[0].endswith(u"汇总") else 1))
            resultList.extend(sortedRows)

        return resultList

    @classmethod
    def getFileId(cls):
        """
        获取文件id
        :return:
        """
        return uuid.uuid4().hex

    @classmethod
    def generateWorkSheet(cls, queryset, wb, retCols, title="", exportType="doctor", enableFR=False, is_need_summary=1):
        """
        数据写入excel
        :return:
        """
        # ws = wb.create_sheet(u"病历统计")
        ws = wb.active
        # ws.title = u""
        colLen = len(retCols) - 1

        titleResult = [retCols[0], retCols[1]]
        if enableFR:
            titleResult.append(retCols[20])
        else:
            colLen = colLen - 1
        colLetter = get_column_letter(colLen)
        mergeStr = "A1:%s1" % colLetter
        ws.merge_cells(mergeStr)
        ws["A1"] = title
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws["A1"].alignment = alignment

        titleResult.extend(retCols[11:20])
        if exportType == "doctor":
            titleResult.extend(retCols[22: 24])
        else:
            titleResult.extend(retCols[20: 21])
        titleResult.extend(retCols[2:11])
        ws.append(tuple(titleResult))
        tmp_list = []
        if exportType == "doctor":
            if enableFR:
                for row_queryset in queryset:
                    r = row_queryset
                    if (r[0], r[1]) in tmp_list:
                        continue
                    tmp_list.append((r[0], r[1]))
                    if r[20] and r[21] and r[20] != "--":
                        tempDoctorFR = "{0}{1}".format(r[20], r[21])
                    else:
                        tempDoctorFR = r[20]
                    tempDoctorFR = tempDoctorFR if tempDoctorFR else ""
                    timelyRate = "{:.2f}%".format(r[22]) if r[22] != "-" else r[22]
                    fixRate = "{:.2f}%".format(r[23]) if r[23] != "-" else r[23]
                    retRow = [r[0], r[1], tempDoctorFR, r[11], r[12], "{:.2f}%".format(r[13]), r[14], r[15],
                            "{:.2f}%".format(r[16]), r[17], r[18], timelyRate, fixRate, r[2], r[3], r[4], r[5], r[6],
                            "{:.2f}%".format(r[7]), "{:.2f}%".format(r[8]), "{:.2f}%".format(r[9]), r[10]]
                    ws.append(tuple(retRow))
            else:
                for row_queryset in queryset:
                    r = row_queryset
                    if (r[0], r[1]) in tmp_list:
                        continue
                    tmp_list.append((r[0], r[1]))
                    timelyRate = "{:.2f}%".format(float(r[22])) if r[22] != "-" else r[22]
                    fixRate = "{:.2f}%".format(float(r[23])) if r[23] != "-" else r[23]
                    retRow = [r[0], r[1], r[11], r[12], "{:.2f}%".format(r[13]), r[14], r[15], "{:.2f}%".format(r[16]),
                              r[17], r[18], "{:.2f}%".format(r[19]), timelyRate, fixRate, r[2], r[3], r[4], r[5], r[6],
                              "{:.2f}%".format(r[7]), "{:.2f}%".format(r[8]), "{:.2f}%".format(r[9]), r[10]]
                    ws.append(tuple(retRow))
        else:
            for row_queryset in queryset:
                r = row_queryset
                if is_need_summary != 1 and "汇总" in (r[0] or ""):
                    continue
                if (r[0], r[1]) in tmp_list:
                    continue
                tmp_list.append((r[0], r[1]))
                # print "row type is : %s" % type(row_queryset[1:-1])
                rowData = list(row_queryset[:-1])
                rowData[6] = "{:.2f}%".format(rowData[6])
                rowData[7] = "{:.2f}%".format(rowData[7])
                rowData[8] = "{:.2f}%".format(rowData[8])
                rowData[12] = "{:.2f}%".format(rowData[12])
                rowData[15] = "{:.2f}%".format(rowData[15])
                rowData[18] = "{:.2f}%".format(rowData[18])
                rowData[19] = "{:.2f}%".format(float(rowData[19])) if rowData[19] != "-" else rowData[19]
                rowData[20] = "{:.2f}%".format(float(rowData[20])) if rowData[20] != "-" else rowData[20]

                tmp = rowData[:2] + rowData[11: 21] + rowData[2: 11]
                ws.append(tuple(tmp))

    @classmethod
    def format_header_data(cls, response, header, data, is_sort=0, not_sort_header=None, first_title_dict={}):
        """
        序列化标题+数据
        """
        if not_sort_header is None:
            not_sort_header = []
        for item in header:
            if item.name in ["sort_time"]:
                continue
            protoHeader = {}  # response.headers.add()
            protoHeader["title"] = item.name if not hasattr(item, "displayName") else item.displayName
            protoHeader["key"] = item.name
            if "-" in item.name:
                protoHeader["firstTitle"] = item.name.split("-")[0]
            if first_title_dict:
                protoHeader["firstTitle"] = first_title_dict.get(item.name, "")
            if item.name not in not_sort_header and is_sort:
                protoHeader["isSort"] = is_sort
            response["headers"].append(protoHeader)
        for row_data in data:
            if StatsRepository.verify_row_data(row_data):
                protoData = {"item": {}}  # response.data.add()
                for key, value in row_data.items():
                    protoData["item"][key] = str(value)
                response["data"].append(protoData)

    @classmethod
    def getRequestStartAndSize(cls, request):
        """
        获取分页请求中的start和size
        :return:
        """
        size = 10
        start = 0
        MAXSIZE = 1000
        if request.size and 0 < request.size <= MAXSIZE:
            size = request.size
        if request.size > MAXSIZE:
            size = MAXSIZE
        if request.start:
            start = request.start
        return start, size

    @classmethod
    def sortArchivingRateResult(cls, queryset, sortKey, sortWay):
        """
        医生归档率结果排序
        :return:
        """
        resultList = [queryset[0]]
        deptDict = defaultdict(list)
        for item in queryset[1:]:
            deptDict[item[0]].append(item)

        deptList = list(deptDict.items())
        try:
            if sortWay == "desc":
                sortedDept = sorted(deptList, key=lambda x: cls.getDeptsSortVal(x, deptDict)[sortKey] or 0, reverse=True)
            else:
                sortedDept = sorted(deptList, key=lambda x: cls.getDeptsSortVal(x, deptDict)[sortKey] or 0)
        except TypeError:
            logging.error("sortArchivingRateResult, error: %s", traceback.format_exc())
            sortedDept = deptList
        for item in sortedDept:
            if sortWay == "desc":
                # x[20]是指F/R
                sortedRows = sorted(item[1] or 0,
                                    key=lambda x: (2 if (x[1] or "-") == "--" else 1, x[1] or "-", 2 if (x[20] or "-") == "--" else 1, x[sortKey]or 0),
                                    reverse=True)
            else:
                sortedRows = sorted(item[1] or 0,
                                    key=lambda x: (0 if (x[1] or "-") == "--" else 1, x[1] or "-", 0 if (x[20] or "-") == "--" else 1, x[sortKey] or 0))
            resultList.extend(sortedRows)

        return resultList

    @classmethod
    def getDeptsSortVal(cls, deptsListItem, deptDict):
        """
        从sortArchivingRateResult 的deptsList 获取外层科室排序的对应值
        :return:
        """
        dept = deptsListItem[0]
        result = deptDict[dept][0] or "-"
        flag = False
        for item in deptDict[dept]:
            if item[1] == "--":
                result = item
                flag = True
                break
        if not flag:
            logging.info("department %s has no total stats!" % dept)

        return result

    @classmethod
    def sortDirectorStatsResult(cls, queryset, sortKey, sortWay="asc"):
        """
        科主任统计结果排序
        :return:
        """
        resultList = [queryset[0]]
        deptDict = defaultdict(list)
        for item in queryset[1:]:
            deptDict[item[0]].append(item)

        deptList = list(deptDict.items())

        if sortWay == "desc":
            sortedDept = sorted(deptList, key=lambda x: cls.getDeptsSortVal(x, deptDict)[sortKey], reverse=True)
        else:
            sortedDept = sorted(deptList, key=lambda x: cls.getDeptsSortVal(x, deptDict)[sortKey])
        for item in sortedDept:
            if sortWay == "desc":
                sortedRows = sorted(item[1] or "-", key=lambda x: (2 if (x[1] or "-") == "--" else 1, x[1] or "-"), reverse=True)
            else:
                sortedRows = sorted(item[1] or "-", key=lambda x: (0 if (x[1] or "-") == "--" else 1, x[1] or "-"))
            resultList.extend(sortedRows)

        return resultList

    @classmethod
    def query_docror_archiving_data(cls, app, request):
        """
        查询医生(科主任)归档率明细数据
        :return:
        """
        start, size = cls.getRequestStartAndSize(request)
        conditionStr = ""
        params = []
        # 构造查询条件
        if request.startTime:
            conditionStr += (" and " if conditionStr != "" else "") + " c.dischargeTime >= '%s' "
            params.append(request.startTime)
        # 出院结束时间
        if request.endTime:
            conditionStr += (" and " if conditionStr != "" else "") + " c.dischargeTime <= '%s' "
            if len(request.endTime) <= 10:
                params.append(request.endTime + " 23:59:59")
            else:
                params.append(request.endTime)
        if request.branch:
            conditionStr += (" and " if conditionStr != "" else "") + " c.branch = '%s' "
            params.append(request.branch)
        if request.department:
            deptList = request.department.split(",")
            formatStr = ",".join(['"%s"' % item for item in deptList])
            conditionStr += (" and " if conditionStr != "" else "") + " c.outDeptName in (%s)" % formatStr
        if request.doctor and request.doctor != "--":
            doctorList = request.doctor.split(",")
            doctorListStr = ",".join(['"%s"' % item for item in doctorList])
            conditionStr += (" and " if conditionStr != "" else "") + " c.attendDoctor in (%s) " % doctorListStr
        if hasattr(request, "applyFlag") and request.applyFlag:
            conditionStr += (" and " if conditionStr != "" else "") + " cx.applyflag = '%s' "
            params.append(request.applyFlag)
        if hasattr(request, "reviewFlag") and request.reviewFlag:
            conditionStr += (" and " if conditionStr != "" else "") + " cx.auditflag = '%s' "
            params.append(request.reviewFlag)
        if hasattr(request, "isPrimaryDiagValid") and request.isPrimaryDiagValid:
            conditionStr += (" and " if conditionStr != "" else "") + " cx.isprimarydiagvalid = '%s' "
            params.append(request.isPrimaryDiagValid)
        if hasattr(request, "isMinorDiagValid") and request.isMinorDiagValid:
            conditionStr += (" and " if conditionStr != "" else "") + " cx.isminordiagvalid = '%s' "
            params.append(request.isMinorDiagValid)
        if hasattr(request, "isPrimaryOperValid") and request.isPrimaryOperValid:
            conditionStr += (" and " if conditionStr != "" else "") + " cx.isprimaryopervalid = '%s' "
            params.append(request.isPrimaryOperValid)
        if hasattr(request, "isMinorOperValid") and request.isMinorOperValid:
            conditionStr += (" and " if conditionStr != "" else "") + " cx.isminoropervalid = '%s' "
            params.append(request.isMinorOperValid)
        if hasattr(request, "doctorFR") and request.doctorFR:
            if request.doctorFRFlag == "F":
                conditionStr += (" and " if conditionStr != "" else "") + " cx.fellowdoctor = '%s' "
                params.append(request.doctorFR)
            elif request.doctorFRFlag == "R":
                conditionStr += (" and " if conditionStr != "" else "") + " cx.residentdoctor = '%s' "
                params.append(request.doctorFR)
            else:
                conditionStr += (" and " if conditionStr != "" else "") \
                                + " ( FIND_IN_SET(cx.residentdoctor, '%s') or FIND_IN_SET(cx.fellowdoctor, '%s')) "
                params.append(request.doctorFR)
                params.append(request.doctorFR)

        if conditionStr:
            conditionStr = " where " + conditionStr

        # 排序规则
        if request.sortKey:
            if request.sortKey not in ["inpDays", "dischargeTime", "caseId", "admitTime"]:
                # self.logger.error("sortKey is not support.")
                return None, None, None
        if request.sortWay:
            if request.sortWay.lower() not in ["asc", "desc", "null"]:
                # self.logger.error("sortWay is not support.")
                return None, None, None
        sort = (request.sortKey if request.sortKey else "dischargeTime") + " " + (
            request.sortWay if request.sortWay and request.sortWay != "null" else "ASC")

        # get case list,一共30个字段
        with app.mysqlConnection.session() as cursor:
            # get total count
            query_count_sql = """select count(id) from `case` as c join case_extend cx on c.caseId = cx.caseid 
                    left join dim_dept_statis d on c.outDeptId = d.deptid """ + (conditionStr if conditionStr else "")
            query_count_sql = query_count_sql % tuple(params)
            # self.logger.info("GetDoctorArchivingRateCase, query_count_sql: %s", query_count_sql)
            query = cursor.execute(query_count_sql)
            ret = query.fetchone()
            totalCount = ret[0]

            query_data_sql = """select id, c.caseId,c.patientId, c.name, c.gender, c.hospital, c.branch,
                          c.outDeptName, c.attendDoctor, c.admitTime,c.dischargeTime,cx.outdeptname,
                          c.inpDays, cx.applyflag, c.applytime, c.applyDoctor,cx.auditflag, c.reviewer, c.reviewTime,
                          c.isDead, cx.isSecondArchiving, cx.isThirdArchiving, cx.isSeventhArchiving, c.hasOutDeptProblem ,
                          cx.isprimarydiagvalid, cx.isminordiagvalid, cx.isprimaryopervalid, cx.isminoropervalid
                          from `case` c join case_extend cx on c.caseId = cx.caseid {}
                          order by {} limit %s, %s""".format(conditionStr, sort)
            # left join dim_dept_statis d on c.outDeptId = d.deptid

            params.append(start)
            params.append(size)
            query_data_sql = query_data_sql % tuple(params)
            # self.logger.info("GetDoctorArchivingRateCase, query_data_sql: %s", query_data_sql)
            query = cursor.execute(query_data_sql)
            queryset = query.fetchall()
            # self.logger.info("total len is : %s" % totalCount)
            # self.logger.info("queryset len is : %s" % len(queryset))

        return queryset, totalCount, start

    @classmethod
    def generateDirectorWorkSheet(cls, queryset, wb, retCols, title=""):
        """
        科主任统计导出写入excel
        :return:
        """
        ws = wb.active
        colLen = len(retCols)
        colLetter = get_column_letter(colLen)
        mergeStr = "A1:%s1" % colLetter
        ws.merge_cells(mergeStr)
        ws["A1"] = title
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws["A1"].alignment = alignment
        ws.append(tuple(retCols))
        for row_queryset in queryset:
            rowData = list(row_queryset[:-1])
            rowData[3] = "{:.2f}%".format(rowData[3])
            rowData[4] = "{:.2f}%".format(rowData[4])
            rowData[5] = "{:.2f}%".format(rowData[5])
            rowData[6] = "{:.2f}%".format(rowData[6])
            ws.append(tuple(rowData))

    @classmethod
    def sortMedicalIndicatorResult(cls, queryset, sortKey, sortWay, removeAllFlag):
        """
        病案指标结果排序
        :return:
        """
        resultList = []
        toSortList = list(queryset)
        if removeAllFlag:
            # 删除“全部”行
            toSortList = toSortList[1:]
        deptDict = defaultdict(list)
        for item in toSortList:
            if item[0]:
                # branch不为空的行不参与排序
                resultList.append(item)
            else:
                deptDict[item[1]].append(item)

        deptList = list(deptDict.items())
        if sortWay == "desc":
            sortedDept = sorted(deptList, key=lambda x: x[1][0][sortKey], reverse=True)
        else:
            sortedDept = sorted(deptList, key=lambda x: x[1][0][sortKey])
        for item in sortedDept:
            if sortWay == "desc":
                sortedRows = sorted(item[1] or "-", key=lambda x: (2 if (x[2] or "-") == "--" else 1, x[2] or "-"), reverse=True)
            else:
                sortedRows = sorted(item[1] or "-", key=lambda x: (0 if (x[2] or "-") == "--" else 1, x[2] or "-"))
            resultList.extend(sortedRows)

        return resultList

    @classmethod
    def generateMedicalIndicatorWorkSheet(cls, queryset, wb, retCols, title=""):
        """
        病案指标统计导出写入excel
        :return:
        """
        ws = wb.active
        ws.append(tuple(retCols))
        for row_queryset in queryset:
            rowData = list(row_queryset[:-1])
            ws.append(tuple(rowData))

    @classmethod
    def query_medical_indicator_data(cls, app, request):
        """
        查询病案指标明细数据
        :return:
        """
        start, size = cls.getRequestStartAndSize(request)
        conditionStr = ""
        params = []
        # 构造查询条件
        if request.startTime:
            conditionStr += (" and " if conditionStr != "" else "") + " c.dischargeTime >= '%s' "
            params.append(request.startTime)
        # 出院结束时间
        if request.endTime:
            conditionStr += (" and " if conditionStr != "" else "") + " c.dischargeTime <= '%s' "
            if len(request.endTime) <= 10:
                params.append(request.endTime + " 23:59:59")
            else:
                params.append(request.endTime)
        if request.branch:
            conditionStr += (" and " if conditionStr != "" else "") + " c.branch = '%s' "
            params.append(request.branch)
        if request.department:
            deptList = request.department.split(",")
            formatStr = ','.join(['"%s"' % item for item in deptList])
            conditionStr += (" and " if conditionStr != "" else "") \
                            + " ifnull(d.statis_name, c.outDeptName) in (%s)" % formatStr
        if request.doctor and request.doctor != "--":
            doctorList = request.doctor.split(",")
            doctorListStr = ",".join(['%s'] * len(doctorList))
            conditionStr += (" and " if conditionStr != "" else "") + " c.attendDoctor in (%s) " % doctorListStr
        if request.status:
            conditionStr += (" and " if conditionStr != "" else "") + " c.status = '%s' "
            params.append(request.status)

        if conditionStr:
            conditionStr = " where " + conditionStr

        # 排序规则
        if request.sortKey:
            if request.sortKey not in ["dischargeTime", "admitTime"]:
                # self.logger.error("sortKey is not support.")
                return None, None, None
        if request.sortWay:
            if request.sortWay.lower() not in ["asc", "desc", "null"]:
                # self.logger.error("sortWay is not support.")
                return None, None, None
        sort = (request.sortKey if request.sortKey else "dischargeTime") + " " + (
            request.sortWay if request.sortWay and request.sortWay != "null" else "ASC")

        # get case list
        with app.mysqlConnection.session() as cursor:
            # get total count
            query_count_sql = """select count(id) from `case` as c join case_extend cx on c.caseId = cx.caseid 
                        left join dim_dept_statis d on c.outDeptId = d.deptid """ + (
                conditionStr if conditionStr else "")
            query_count_sql = query_count_sql % tuple(params)
            # self.logger.info("GetMedicalIndicatorStatsCase, query_count_sql: %s", query_count_sql)
            query = cursor.execute(query_count_sql)
            ret = query.fetchone()
            totalCount = ret[0]

            query_data_sql = """select id, c.caseId,c.patientId, c.name, c.gender, c.hospital, c.branch,
                              c.outDeptName, c.attendDoctor, c.admitTime,c.dischargeTime,cx.outdeptname, 
                              c.status,cx.inrecord24hcomplete, cx.opsrecord24hcomplete, cx.outrecord24hcomplete, 
                              cx.firstpage24hcomplete, cx.opsrecordwhole, cx.doctorroundswhole, 
                              cx.rescuerecord6hcomplete, cx.isSecondArchiving,
                              (case when cx.auditflag='已退回' then '否' else '是' end ) ,cx.isemrcopy ,cx.isnormmrc 
                              from `case` c join case_extend cx on c.caseId = cx.caseid 
                              left join dim_dept_statis d on c.outDeptId = d.deptid {} 
                              order by {} limit %s, %s""".format(conditionStr, sort)
            params.append(start)
            params.append(size)
            query_data_sql = query_data_sql % tuple(params)
            # self.logger.info("GetMedicalIndicatorStatsCase, query_data_sql: %s", query_data_sql)
            query = cursor.execute(query_data_sql)
            queryset = query.fetchall()
            # self.logger.info("total len is : %s" % totalCount)
            # self.logger.info("queryset len is : %s" % len(queryset))

        return queryset, totalCount, start

    @classmethod
    def getRequestPageAndCount(cls, request):
        """
        获取分页请求中的start和size
        :param request:
        :return:
        """
        size = 1000
        start = 0
        MAXSIZE = 1000
        if request.count and 0 < request.count <= MAXSIZE:
            size = request.count
        if request.count > MAXSIZE:
            size = MAXSIZE

        if request.page:
            start = (request.page - 1) * request.count
        return start, size
    
    @classmethod
    def get_call_proc_sql(cls, target, request, sixMonth=0):
        """
        格式化存储过程
        :return:
        """
        case_type_dict = {"running": "运行病历", "archived": "归档病历", "final": "终末病历"}
        dept_type_dict = {1: "内科", 2: "外科", 0: "全部"}
        stats_type_dict = {1: "qctype", 2: "deptkind", 3: "area"}
        deptType = dept_type_dict[request.deptType]
        caseType = case_type_dict.get(request.caseType, "全部类型")
        department = request.department or "" if target != "医生成绩统计" else request.department or "全部科室"
        doctorName = request.doctorName or "" if target != "医生成绩统计" else request.doctorName or "全部医生"
        startTime = request.startTime
        endTime = request.endTime
        new_start_time = StatsRepository.get_start_time(endTime)
        if sixMonth:
            startTime = new_start_time
            endTime = StatsRepository.get_end_time(endTime)
        call_proc_sql = """call pr_case_expert('{target}','{caseType}','{startTime}','{endTime}','{branch}',
        '{deptType}','{department}','{doctor}','','{statsType}')""".format(
            target=target, caseType=caseType, startTime=startTime,
            endTime=endTime, branch=request.branch or "全部院区", deptType=deptType,
            department=department, doctor=doctorName, statsType=stats_type_dict[request.statsType or 3])
        return call_proc_sql

    @classmethod
    def get_common_stats_pic_response(cls, app, call_proc_sql, response, request, level=2, is_rate=0):
        """
        专家统计通用获取 折线/柱状图数据
        :return:
        """
        endTime = request.endTime
        startTime = StatsRepository.get_start_time(endTime)
        endTime = StatsRepository.get_end_time(endTime)
        with app.mysqlConnection.session() as cursor:
            query = cursor.execute(call_proc_sql)
            # self.logger.info("get_common_stats_pic_response, call_proc_sql: %s", call_proc_sql)
            queryset = query.fetchall()
            res_dict = {}
            if queryset:
                for item in queryset:
                    if item[0] not in res_dict:
                        res_dict[item[0]] = StatsRepository.get_init_dict(startTime, endTime)
                    data = item[level]
                    if is_rate and level != 3:
                        data = item[level] / item[2] * 100
                    res_dict[item[0]][item[1]] = data
                for branch in res_dict:
                    if not branch:
                        continue
                    protoItem = {"items": []}  # response.data.add()
                    protoItem["dataName"] = branch
                    detail_data = res_dict[branch]
                    for x, y in detail_data.items():
                        detailData = {}  # protoItem.items.add()
                        detailData["xData"] = str(x)
                        detailData["yData"] = cls.keep_one(y)
                        protoItem["items"].append(detailData)
                    response["data"].append(protoItem)

    @classmethod
    def keep_one(cls, num, ratio=False):
        """
        保留一位小数
        :param num:
        :param ratio: 是否带单位符号
        :return:
        """
        if not num:
            num = 0
        minus = False
        if num < 0:
            num = float(str(num)[1:])
            minus = True
        symbol = '%'
        if num > 100:
            x = '%.1f' % num
        elif num >= 1:
            x = '%0.3g' % num  # '1.xx or 11.x'
        elif num >= 0.1:
            x = '%0.2f' % num  # 0.xx
        elif num >= 0.01:
            if ratio:
                n = num * 10
                x, symbol = '%0.2f' % n, '‰'  # 0.xx
            else:
                x = '%.1f' % num
        elif num > 0.0001:
            if ratio:
                n = num * 100
                x, symbol = '%0.2f' % n, '‱'
            else:
                x = '%.1f' % num
        else:
            x = '0'
        if minus and x != '0':
            x = "-" + x
        if ratio:
            return x + symbol
        return x

    @classmethod
    def get_common_header_data(cls, app, target_name, request):
        """
        专家统计 通用获取标题+数据列表数据+配置
        :return:
        """
        qc_stats_level = app.config.get(Config.QC_STATS_LEVEL)
        ward_field = "病区"
        if app.config.get(Config.QC_ASSIGN_DIMENSION) == 1:
            ward_field = "科室"
        qc_stats_level = qc_stats_level.split(",")
        total = 0
        with app.mysqlConnection.session() as cursor:
            data = []
            if target_name == "质控情况分布":
                call_proc_sql = cls.get_call_proc_sql(target_name, request)
                query = cursor.execute(call_proc_sql)
                # self.logger.info("get_common_stats_pic_response, call_proc_sql: %s", call_proc_sql)
                queryset = query.fetchall()
                data_yaml = EXPERT_ALL_SCORE_YAML.format(first=qc_stats_level[0], second=qc_stats_level[1],
                                                         third=qc_stats_level[2])
                for item in queryset:
                    tmp = {"院区": item[0] or "", "平均分": cls.keep_one(item[2]), "总数": item[1] or 0,
                           "{}率".format(qc_stats_level[0]): cls.keep_one(item[6], True),
                           "{}率".format(qc_stats_level[1]): cls.keep_one(item[7], True),
                           "{}率".format(qc_stats_level[2]): cls.keep_one(item[8], True),
                           "{}数".format(qc_stats_level[0]): item[3], "{}数".format(qc_stats_level[1]): item[4],
                           "{}数".format(qc_stats_level[2]): item[5]}
                    data.append(tmp)
                total = len(data)
            elif target_name == "科室成绩概览":
                call_proc_sql = cls.get_call_proc_sql(target_name, request)
                query = cursor.execute(call_proc_sql)
                # self.logger.info("get_common_stats_pic_response, call_proc_sql: %s", call_proc_sql)
                queryset = query.fetchall()
                data_yaml = INTERNAL_SURGERY_LIST_YAML.replace("[first]", qc_stats_level[0]).replace(
                    "[second]", qc_stats_level[1]).replace("[third]", qc_stats_level[2]).replace("[ward]", ward_field)
                for item in queryset:
                    if not item[0]:
                        continue
                    tmp = {"内外科": item[0], ward_field: item[1], "总数": int(item[2]), "总分": float(item[3]),
                           "{}数".format(qc_stats_level[0]): int(item[4]), "{}数".format(qc_stats_level[1]): int(item[5]),
                           "{}数".format(qc_stats_level[2]): int(item[6])}
                    data.append(tmp)
                total = len(data)
            elif target_name == "医生成绩统计":
                call_proc_sql = cls.get_call_proc_sql(target_name, request)
                query = cursor.execute(call_proc_sql)
                # self.logger.info("get_common_stats_pic_response, call_proc_sql: %s", call_proc_sql)
                queryset = query.fetchall()
                data_yaml = DOCTOR_SCORE_LIST_YAML.replace("[first]", qc_stats_level[0]).replace(
                    "[second]", qc_stats_level[1]).replace("[third]", qc_stats_level[2]).replace("[ward]", ward_field)
                item_1_list = []
                item_0_list = []
                for item in queryset:
                    if item[1] and item[1] not in item_1_list:
                        item_1_list.append(item[1])
                    if item[0] not in item_0_list:
                        item_0_list.append(item[0])
                    tmp = {ward_field: item[0] or "", "诊疗小组": item[1] or "", "医生姓名": item[2] or "",
                           "总分": float(item[4]), "总数": int(item[3]),
                           "{}数".format(qc_stats_level[0]): int(item[5]), "{}数".format(qc_stats_level[1]): int(item[6]),
                           "{}数".format(qc_stats_level[2]): int(item[7])}
                    data.append(tmp)
                total = len(data) + len(item_1_list) + len(item_0_list) + 1

            return data, data_yaml, total

    @classmethod
    def get_file_id(cls, file_name):
        """
        获取文件id
        :return:
        """
        return uuid.uuid3(uuid.NAMESPACE_DNS, file_name + str(random.randint(100, 1000))).hex
    
    @classmethod
    def getLastUpdateTime(cls, app, u_type):
        """
        查询上次更新时间
        :return:
        """
        result = {}
        with app.mysqlConnection.session() as cursor:
            table = cls.get_update_status_table(u_type)
            query_last_update_time_sql = "select updatetime, updatestatus from %s" % table
            query = cursor.execute(query_last_update_time_sql)
            # self.logger.info("getLastUpdateTime, query_last_update_time_sql: %s", query_last_update_time_sql)
            ret = query.fetchone()
            if ret:
                result["status"] = int(ret[1] or 0)
                updateDatetime = ret[0] + timedelta(hours=8) if ret[0] else ""
                # self.logger.info("getLastUpdateTime, updateDatetime: %s", updateDatetime)
                result["updateTime"] = updateDatetime
        return result

    @classmethod
    def get_defect_rate_data(cls, app, request, is_need_group=1):
        """
        获取缺陷率统计数据
        :return:
        """
        dept = request.department or "全部"
        if dept == "总计":
            dept = "全部"
        ward = request.ward or "全部"
        if ward == "总计":
            ward = "全部"
        args = ["{}缺陷率统计".format(request.type), request.startTime, request.endTime, request.branch or "全部院区",
                dept, ward, request.group or "全部", request.attend or "全部"]
        with app.mysqlConnection.session() as session:
            call_sql = "call pr_case_rate('%s','%s','%s','%s','%s','%s','%s','%s');" % tuple(args)
            # self.logger.info("StatsDefectRateList, call_sql: %s", call_sql)
            query = session.execute(call_sql)
            queryset = query.fetchall()
            data = StatsRepository.format_defect_rate_data(queryset, request, is_need_group)
        return data
